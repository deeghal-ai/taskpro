# projects/report_views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from datetime import date, timedelta, datetime
from accounts.models import User
from .services import ReportingService, TATAnalyticsService, AgeingReportService, ProjectService, GeneralReportService
from .views import ensure_has_management_access
from .models import Project, ProjectDelivery, TaskAssignment, Product
from locations.models import City
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import json

# Import permission helpers from views
def ensure_has_management_access(request):
    """Check if user has management access (DPM, VIDEO_PM, or SENIOR_MANAGER)"""
    if request.user.role not in ['DPM', 'VIDEO_PM', 'SENIOR_MANAGER']:
        messages.error(request, "You don't have permission to access this page.")
        return redirect('projects:project_list')
    return None

def ensure_has_full_management_access(request):
    """Check if user has full management access (DPM or VIDEO_PM only)"""
    if request.user.role not in ['DPM', 'VIDEO_PM']:
        messages.error(request, "You don't have permission to perform this action.")
        return redirect('projects:project_list')
    return None


@login_required
def team_member_report(request, team_member_id=None):
    """View for team member productivity report - now using on-demand calculations"""
    # Check permissions - allow TEAM_MEMBER to view own report, or management roles to view any
    if team_member_id:
        # Viewing another user's report - requires management access
        redirect_response = ensure_has_management_access(request)
        if redirect_response:
            return redirect_response
        team_member = get_object_or_404(User, id=team_member_id)
    elif request.user.role == 'TEAM_MEMBER':
        # Team member viewing own report
        team_member = request.user
    else:
        # Management user without specific team_member_id - redirect to team overview
        return redirect('projects:team_overview_report')
    
    # Get date range from request or default to last 30 days
    end_date = request.GET.get('end_date', date.today())
    if isinstance(end_date, str):
        end_date = date.fromisoformat(end_date)
    
    start_date = request.GET.get('start_date', end_date - timedelta(days=30))
    if isinstance(start_date, str):
        start_date = date.fromisoformat(start_date)
    
    # Get report data (now calculated on-demand!)
    metrics = ReportingService.get_team_member_metrics(team_member, start_date, end_date)
    
    # Get delivery history for details
    delivery_history = ProjectDelivery.objects.filter(
        project_incharge=team_member,
        delivery_date__range=[start_date, end_date]
    ).select_related('project').order_by('-delivery_date')
    
    # Prepare context in the format expected by template
    report_data = {
        'period': metrics['period'],
        'delivery_history': delivery_history,
        'summary': {
            'average_productivity': metrics['productivity']['score'],
            'average_optimization': metrics['optimization']['score'],
            'average_utilization': metrics['utilization']['score'],
            'average_efficiency': metrics['efficiency']['score'],
            'average_quality_rating': metrics['quality']['average_rating'],
            'average_delivery_rating': metrics['delivery']['average_rating'],
            'total_assignments_completed': metrics['quality']['total_assignments'],
            'total_projects_delivered': metrics['delivery']['total_projects'],
            'on_time_delivery_rate': metrics['delivery']['on_time_rate'],
            'total_hours_projected': metrics['productivity']['projected_hours'],
            'total_hours_worked': metrics['productivity']['worked_hours'],
            'optimization_saved_hours': metrics['optimization']['saved_hours'],
            'efficiency_total_work_hours': metrics['efficiency']['total_work_minutes'] / 60 if metrics['efficiency']['total_work_minutes'] else 0,
            'efficiency_misc_hours': metrics['efficiency']['misc_minutes'] / 60 if metrics['efficiency']['misc_minutes'] else 0,
        }
    }
    
    context = {
        'team_member': team_member,
        'report': report_data,
        'start_date': start_date,
        'end_date': end_date,
        'title': f'Productivity Report - {team_member.get_full_name()}'
    }
    
    return render(request, 'projects/reports/team_member_report.html', context)

@login_required
def team_overview_report(request):
    """Overview report for all team members (Management view) - now using on-demand calculations"""
    # Check if user has management access (includes Senior Managers)
    redirect_response = ensure_has_management_access(request)
    if redirect_response:
        return redirect_response
    
    # Get date range
    end_date = request.GET.get('end_date', date.today())
    if isinstance(end_date, str):
        end_date = date.fromisoformat(end_date)
    
    start_date = request.GET.get('start_date', end_date - timedelta(days=30))
    if isinstance(start_date, str):
        start_date = date.fromisoformat(start_date)
    
    # Get overview data (much simpler now!)
    overview_data = ReportingService.get_team_overview(start_date, end_date)
    
    # Transform data format to match template expectations
    formatted_overview = []
    for item in overview_data:
        formatted_overview.append({
            'team_member': item['team_member'],
            'metrics': {
                'avg_productivity': item['metrics']['productivity']['score'],
                'avg_optimization': item['metrics']['optimization']['score'],
                'avg_utilization': item['metrics']['utilization']['score'],
                'avg_efficiency': item['metrics']['efficiency']['score'],
                'avg_quality': item['metrics']['quality']['average_rating'],
                'avg_delivery': item['metrics']['delivery']['average_rating'],
                'total_assignments': item['metrics']['quality']['total_assignments'],
                'total_projects': item['metrics']['delivery']['total_projects']
            }
        })
    
    # Calculate team averages and totals
    team_averages = {}
    team_totals = {'total_assignments': 0, 'total_projects': 0}
    
    if formatted_overview:
        # Calculate averages for each metric
        metrics_to_average = ['avg_productivity', 'avg_optimization', 'avg_utilization', 'avg_efficiency', 'avg_quality', 'avg_delivery']
        for metric in metrics_to_average:
            values = [item['metrics'][metric] for item in formatted_overview if item['metrics'][metric] is not None]
            team_averages[metric] = sum(values) / len(values) if values else None
        
        # Calculate totals
        team_totals['total_assignments'] = sum(item['metrics']['total_assignments'] or 0 for item in formatted_overview)
        team_totals['total_projects'] = sum(item['metrics']['total_projects'] or 0 for item in formatted_overview)
    
    context = {
        'overview_data': formatted_overview,
        'team_averages': team_averages,
        'team_totals': team_totals,
        'start_date': start_date,
        'end_date': end_date,
        'title': 'Team Overview Report'
    }
    
    return render(request, 'projects/reports/team_overview.html', context)

@login_required
def delivery_performance_report(request):
    """Delivery performance report for project incharges - simplified"""
    # Check if user has management access (includes Senior Managers)
    redirect_response = ensure_has_management_access(request)
    if redirect_response:
        return redirect_response
    
    # Get date range
    end_date = request.GET.get('end_date', date.today())
    if isinstance(end_date, str):
        end_date = date.fromisoformat(end_date)
    
    start_date = request.GET.get('start_date', end_date - timedelta(days=90))
    if isinstance(start_date, str):
        start_date = date.fromisoformat(start_date)
    
    # Get delivery data - much simpler without complex stored metrics
    deliveries = ProjectDelivery.objects.filter(
        delivery_date__range=[start_date, end_date]
    ).select_related('project', 'project_incharge')
    
    # Group by project incharge
    incharge_data = {}
    for delivery in deliveries:
        incharge = delivery.project_incharge
        if incharge not in incharge_data:
            incharge_data[incharge] = {
                'deliveries': [],
                'total': 0,
                'on_time': 0,
                'rating_sum': 0,
                'rating_count': 0
            }
        
        incharge_data[incharge]['deliveries'].append(delivery)
        incharge_data[incharge]['total'] += 1
        
        # Calculate on-time using the actual database fields, not the property
        if (delivery.expected_completion_date and 
            delivery.actual_completion_date <= delivery.expected_completion_date):
            incharge_data[incharge]['on_time'] += 1
        
        if delivery.delivery_performance_rating:
            incharge_data[incharge]['rating_sum'] += delivery.delivery_performance_rating
            incharge_data[incharge]['rating_count'] += 1
    
    # Calculate averages
    report_data = []
    for incharge, data in incharge_data.items():
        avg_rating = (data['rating_sum'] / data['rating_count']) if data['rating_count'] > 0 else None
        on_time_rate = (data['on_time'] / data['total'] * 100) if data['total'] > 0 else 0
        
        report_data.append({
            'team_member': incharge,
            'average_rating': avg_rating,
            'total_deliveries': data['total'],
            'on_time_rate': on_time_rate,
            'recent_deliveries': sorted(data['deliveries'], key=lambda x: x.delivery_date, reverse=True)[:5]
        })
    
    # Sort by average rating
    report_data.sort(key=lambda x: x['average_rating'] or 0, reverse=True)
    
    context = {
        'report_data': report_data,
        'start_date': start_date,
        'end_date': end_date,
        'title': 'Delivery Performance Report'
    }

    return render(request, 'projects/reports/delivery_performance.html', context)

@login_required
def lol_report(request):
    """LoL report with date range filter and team member selection"""
    # Check if user has management access (includes Senior Managers)
    redirect_response = ensure_has_management_access(request)
    if redirect_response:
        return redirect_response
    
    # Initialize form data
    selected_team_members = []
    report_data = []
    
    # Get date range from request or default to last 30 days
    end_date = request.GET.get('end_date', date.today())
    if isinstance(end_date, str):
        end_date = date.fromisoformat(end_date)
    
    start_date = request.GET.get('start_date', end_date - timedelta(days=30))
    if isinstance(start_date, str):
        start_date = date.fromisoformat(start_date)
    
    # Get selected team members from POST or GET
    if request.method == 'POST':
        selected_member_ids = request.POST.getlist('team_members')
    else:
        selected_member_ids = request.GET.getlist('team_members')
    
    # Get all team members for the form
    all_team_members = User.objects.filter(role='TEAM_MEMBER').order_by('first_name', 'last_name')
    
    # If team members are selected, generate the report
    if selected_member_ids:
        selected_team_members = User.objects.filter(
            id__in=selected_member_ids,
            role='TEAM_MEMBER'
        )
        
        # Get report data
        report_data = ReportingService.get_lol_report_data(
            selected_team_members,
            start_date,
            end_date
        )
    
    # Prepare team members with selection state for the form
    team_members_with_selection = []
    for member in all_team_members:
        team_members_with_selection.append({
            'member': member,
            'is_selected': str(member.id) in selected_member_ids
        })
    
    context = {
        'team_members': team_members_with_selection,
        'selected_member_ids': selected_member_ids,
        'selected_team_members': selected_team_members,
        'report_data': report_data,
        'start_date': start_date,
        'end_date': end_date,
        'title': 'LoL Report',
        'has_results': len(report_data) > 0,
    }
    
    return render(request, 'projects/reports/lol_report.html', context)

@login_required
def reports_dashboard(request):
    """Reports dashboard for Senior Managers and management roles"""
    # Check if user has management access (includes Senior Managers)
    redirect_response = ensure_has_management_access(request)
    if redirect_response:
        return redirect_response
    
    # Get some quick stats for the dashboard (optional)
    from accounts.models import User
    from .models import Project, ProjectDelivery
    from datetime import date, timedelta
    
    # Calculate some basic stats
    total_team_members = User.objects.filter(role='TEAM_MEMBER').count()
    
    # Active projects (those not in 'Final Delivery' status)
    active_projects = Project.objects.exclude(
        current_status__category_two='Final Delivery'
    ).count()
    
    # This month's deliveries
    current_month_start = date.today().replace(day=1)
    deliveries_this_month = ProjectDelivery.objects.filter(
        delivery_date__gte=current_month_start
    ).count()
    
    # Calculate average productivity using the same logic as team overview
    end_date = date.today()
    start_date = end_date - timedelta(days=30)  # Last 30 days
    
    # Get team overview data to calculate average productivity
    overview_data = ReportingService.get_team_overview(start_date, end_date)
    
    # Calculate team average productivity
    avg_productivity = None
    if overview_data:
        productivity_scores = [
            item['metrics']['productivity']['score'] 
            for item in overview_data 
            if item['metrics']['productivity']['score'] is not None
        ]
        if productivity_scores:
            avg_productivity = sum(productivity_scores) / len(productivity_scores)
    
    # Get data for status history export modal
    products = Product.objects.filter(is_active=True).order_by('name')
    cities = City.objects.all().order_by('name')
    dpms = User.objects.filter(role__in=['DPM', 'VIDEO_PM']).order_by('first_name', 'last_name')
    
    context = {
        'total_team_members': total_team_members,
        'active_projects': active_projects,
        'deliveries_this_month': deliveries_this_month,
        'avg_productivity': avg_productivity,
        'title': 'Reports Dashboard',
        # Data for status history export modal
        'products': products,
        'cities': cities,
        'dpms': dpms,
    }
    
    return render(request, 'projects/reports/reports_dashboard.html', context)


@login_required
def lol_report_export_excel(request):
    """Enhanced LoL report Excel export with raw metrics and final overview"""
    # Check if user has management access
    redirect_response = ensure_has_management_access(request)
    if redirect_response:
        return redirect_response
    
    # Get parameters from request
    end_date = request.GET.get('end_date', date.today())
    if isinstance(end_date, str):
        end_date = date.fromisoformat(end_date)
    
    start_date = request.GET.get('start_date', end_date - timedelta(days=30))
    if isinstance(start_date, str):
        start_date = date.fromisoformat(start_date)
    
    selected_member_ids = request.GET.getlist('team_members')
    if not selected_member_ids:
        messages.error(request, "No team members selected for export.")
        return redirect('projects:lol_report')
    
    # Get selected team members and report data
    selected_team_members = User.objects.filter(
        id__in=selected_member_ids,
        role='TEAM_MEMBER'
    )
    
    report_data = ReportingService.get_lol_report_data(
        selected_team_members,
        start_date,
        end_date
    )
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LoL Report"
    
    # Define styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    subheader_font = Font(bold=True, size=11)
    subheader_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )
    
    # Set column widths
    ws.column_dimensions['A'].width = 20  # Name
    ws.column_dimensions['B'].width = 12  # Utilization
    ws.column_dimensions['C'].width = 12  # Productivity  
    ws.column_dimensions['D'].width = 15  # Avg Quality Score
    ws.column_dimensions['E'].width = 15  # Sum of Task Hour
    ws.column_dimensions['F'].width = 20  # Sum of Actual Working hour
    ws.column_dimensions['G'].width = 15  # Sum of Working Hours
    ws.column_dimensions['H'].width = 18  # Sum of Projected Hours
    ws.column_dimensions['I'].width = 20  # Individual Ratings
    
    # Overview table columns (starting from column J)
    ws.column_dimensions['J'].width = 20  # Name
    ws.column_dimensions['K'].width = 15  # Project Utilization
    ws.column_dimensions['L'].width = 15  # Productivity
    ws.column_dimensions['M'].width = 15  # Quality Score
    ws.column_dimensions['N'].width = 12  # Quality in %
    ws.column_dimensions['O'].width = 12  # % Total
    ws.column_dimensions['P'].width = 15  # Eligibility Status
    
    # RAW METRICS SECTION (Left side)
    current_row = 1
    
    # Main header for raw metrics
    ws.merge_cells(f'A{current_row}:I{current_row}')
    ws[f'A{current_row}'] = "RAW INDIVIDUAL METRICS"
    ws[f'A{current_row}'].font = Font(bold=True, size=14)
    ws[f'A{current_row}'].alignment = center_alignment
    ws[f'A{current_row}'].fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    current_row += 2
    
    # Raw metrics headers
    raw_headers = [
        "Name", "Utilization %", "Productivity %", "Avg Quality Score", 
        "Sum of Task Hour", "Sum of Actual Working hour", "Sum of Working Hours", "Sum of Projected Hours", "Individual Quality Ratings"
    ]
    
    for col_idx, header in enumerate(raw_headers, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    current_row += 1
    
    # Raw metrics data
    for data in report_data:
        member = data['team_member']
        
        # Get individual quality ratings as comma-separated string
        individual_ratings = ", ".join([f"{rating:.2f}" for rating in data['quality_ratings_list']]) or "No ratings"
        
        row_data = [
            f"{member.first_name} {member.last_name}",
            f"{data['avg_utilization']:.2f}%",
            f"{data['avg_productivity']:.2f}%",
            f"{data['avg_quality_rating']:.2f}" if data['avg_quality_rating'] else "N/A",
            f"{data['utilization_details']['worked_hours']:.1f}",
            f"{data['utilization_details']['available_hours']:.1f}",
            f"{data['productivity_details']['worked_hours']:.1f}",
            f"{data['productivity_details']['projected_hours']:.1f}",
            individual_ratings
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.border = border
            if col_idx > 1:  # Align numbers to center
                cell.alignment = center_alignment
        
        current_row += 1
    
    # FINAL OVERVIEW SECTION (Right side)
    overview_start_row = 1
    overview_col_start = 10  # Column J
    
    # Main header for overview
    ws.merge_cells(f'J{overview_start_row}:P{overview_start_row}')
    ws[f'J{overview_start_row}'] = "FINAL OVERVIEW TABLE"
    ws[f'J{overview_start_row}'].font = Font(bold=True, size=14)
    ws[f'J{overview_start_row}'].alignment = center_alignment
    ws[f'J{overview_start_row}'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    overview_start_row += 2
    
    # Overview headers
    overview_headers = [
        "Name", "Project Utilization - 40% Weightage", "Productivity - 30% Weightage", 
        "Quality Score - 30% Weightage (< number of less mistakes )", "Quality Score in %", 
        "% Total", "Eligibility Status"
    ]
    
    for col_idx, header in enumerate(overview_headers):
        cell = ws.cell(row=overview_start_row, column=overview_col_start + col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    overview_start_row += 1
    
    # Overview data
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red for failing criteria
    
    for data in report_data:
        member = data['team_member']
        
        overview_row_data = [
            f"{member.first_name} {member.last_name}",
            f"{data['avg_utilization']:.2f}%",
            f"{data['avg_productivity']:.2f}%",
            f"{data['avg_quality_rating']:.2f}" if data['avg_quality_rating'] else "N/A",
            f"{data['quality_score']:.2f}%",
            f"{data['total_percentage']:.2f}%",
            data['eligibility_status']
        ]
        
        for col_idx, value in enumerate(overview_row_data):
            cell = ws.cell(row=overview_start_row, column=overview_col_start + col_idx, value=value)
            cell.border = border
            
            # Apply red background only to cells that fail criteria
            if col_idx == 1 and data['avg_utilization'] < 85:  # Utilization column
                cell.fill = red_fill
            elif col_idx == 2 and data['avg_productivity'] < 95:  # Productivity column
                cell.fill = red_fill
            elif col_idx == 3 and (data['avg_quality_rating'] is None or data['avg_quality_rating'] < 2.95):  # Quality rating column
                cell.fill = red_fill
            
            if col_idx > 0:  # Align data to center
                cell.alignment = center_alignment
        
        overview_start_row += 1
    
    # Add legend/criteria at the bottom
    legend_row = max(current_row + 3, overview_start_row + 3)
    
    ws.merge_cells(f'A{legend_row}:P{legend_row}')
    ws[f'A{legend_row}'] = "ELIGIBILITY CRITERIA"
    ws[f'A{legend_row}'].font = Font(bold=True, size=12)
    ws[f'A{legend_row}'].alignment = center_alignment
    legend_row += 1
    
    criteria_text = [
        "• Utilization: Must be ≥ 85%",
        "• Productivity: Must be ≥ 95%", 
        "• Quality Rating: Must be ≥ 2.95"
    ]
    
    for criterion in criteria_text:
        ws[f'A{legend_row}'] = criterion
        ws[f'A{legend_row}'].font = Font(bold=True)
        legend_row += 1
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"lol_report_{start_date}_{end_date}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
def general_report_dashboard(request):
    """
    General Report Dashboard with date range filtering and multiple business metrics.
    Shows Sales Confirmed and other key metrics with quantity weighting.
    """
    # Check permissions - allow SENIOR_MANAGER, DPM, and VIDEO_PM
    redirect_response = ensure_has_management_access(request)
    if redirect_response:
        return redirect_response
    
    # Parse filters from request
    filters = {}
    if request.GET.get('date_from'):
        try:
            from datetime import datetime
            filters['date_from'] = datetime.strptime(request.GET.get('date_from'), '%Y-%m-%d').date()
        except ValueError:
            pass
    else:
        # Default to July 1, 2025 if no date_from provided
        from datetime import date
        filters['date_from'] = date(2025, 7, 1)
    
    if request.GET.get('date_to'):
        try:
            from datetime import datetime
            filters['date_to'] = datetime.strptime(request.GET.get('date_to'), '%Y-%m-%d').date()
        except ValueError:
            pass
    
    if request.GET.get('product'):
        filters['product'] = request.GET.get('product')
    
    if request.GET.get('dpm'):
        filters['dpm'] = request.GET.get('dpm')
    
    if request.GET.get('city'):
        filters['city'] = request.GET.get('city')
    
    # Get general report data from service
    from .services import GeneralReportService
    report_data = GeneralReportService.get_general_report_data(filters)
    
    if not report_data.get('success'):
        messages.error(request, f"Error generating general report: {report_data.get('error', 'Unknown error')}")
        report_data = {
            'sales_confirmed': {
                'total_quantity': 0,
                'project_count': 0,
                'product_breakdown': [],
                'dpm_breakdown': []
            }
        }
    
    # Get filter options for dropdowns
    from .models import Product
    from accounts.models import User
    from locations.models import City
    
    products = Product.objects.all().values_list('name', flat=True).distinct()
    dpms = User.objects.filter(role='DPM').values_list('username', flat=True)
    cities = City.objects.all().values_list('name', flat=True).distinct()
    
    # Prepare chart data for Sales Confirmed - Product breakdown
    sales_confirmed = report_data.get('sales_confirmed', {})
    product_breakdown = sales_confirmed.get('product_breakdown', [])
    product_labels = [item['product_name'] for item in product_breakdown]
    product_quantities = [item['quantity'] for item in product_breakdown]
    
    product_chart_json = json.dumps({
        'labels': product_labels,
        'data': product_quantities,
        'colors': ['#dc3545', '#fd7e14', '#ffc107', '#198754', '#0dcaf0', '#6f42c1', '#d63384', '#20c997']
    })
    
    # Prepare DPM chart data for Sales Confirmed
    dpm_breakdown = sales_confirmed.get('dpm_breakdown', [])
    dpm_labels = [item['dpm_name'] for item in dpm_breakdown]
    dpm_quantities = [item['quantity'] for item in dpm_breakdown]
    
    dmp_chart_json = json.dumps({
        'labels': dpm_labels,
        'data': dpm_quantities,
        'colors': ['#198754', '#0dcaf0', '#6f42c1', '#dc3545', '#fd7e14', '#ffc107']
    })
    
    # Prepare chart data for 1st Cut Deliveries
    first_cut_deliveries = report_data.get('first_cut_deliveries', {})
    fcd_product_breakdown = first_cut_deliveries.get('product_breakdown', [])
    fcd_product_labels = [item['product_name'] for item in fcd_product_breakdown]
    fcd_product_quantities = [item['quantity'] for item in fcd_product_breakdown]
    
    fcd_product_chart_json = json.dumps({
        'labels': fcd_product_labels,
        'data': fcd_product_quantities,
        'colors': ['#6f42c1', '#20c997', '#fd7e14', '#dc3545', '#198754', '#0dcaf0', '#ffc107', '#d63384']
    })
    
    # Prepare DPM chart data for 1st Cut Deliveries
    fcd_dpm_breakdown = first_cut_deliveries.get('dpm_breakdown', [])
    fcd_dpm_labels = [item['dpm_name'] for item in fcd_dpm_breakdown]
    fcd_dpm_quantities = [item['quantity'] for item in fcd_dpm_breakdown]
    
    fcd_dpm_chart_json = json.dumps({
        'labels': fcd_dpm_labels,
        'data': fcd_dpm_quantities,
        'colors': ['#6f42c1', '#20c997', '#fd7e14', '#dc3545', '#198754', '#0dcaf0']
    })
    
    # Prepare chart data for Final Deliveries
    final_deliveries = report_data.get('final_deliveries', {})
    fd_product_breakdown = final_deliveries.get('product_breakdown', [])
    fd_product_labels = [item['product_name'] for item in fd_product_breakdown]
    fd_product_quantities = [item['quantity'] for item in fd_product_breakdown]
    
    fd_product_chart_json = json.dumps({
        'labels': fd_product_labels,
        'data': fd_product_quantities,
        'colors': ['#d63384', '#198754', '#0dcaf0', '#fd7e14', '#6f42c1', '#20c997', '#ffc107', '#dc3545']
    })
    
    # Prepare DPM chart data for Final Deliveries
    fd_dpm_breakdown = final_deliveries.get('dpm_breakdown', [])
    fd_dpm_labels = [item['dpm_name'] for item in fd_dpm_breakdown]
    fd_dpm_quantities = [item['quantity'] for item in fd_dpm_breakdown]
    
    fd_dpm_chart_json = json.dumps({
        'labels': fd_dpm_labels,
        'data': fd_dpm_quantities,
        'colors': ['#d63384', '#198754', '#0dcaf0', '#fd7e14', '#6f42c1', '#20c997']
    })
    
    # Prepare chart data for Man Hours
    man_hours = report_data.get('man_hours', {})
    mh_product_breakdown = man_hours.get('product_breakdown', [])
    mh_product_labels = [item['product_name'] for item in mh_product_breakdown]
    mh_product_hours = [item['hours'] for item in mh_product_breakdown]
    
    mh_product_chart_json = json.dumps({
        'labels': mh_product_labels,
        'data': mh_product_hours,
        'colors': ['#28a745', '#17a2b8', '#ffc107', '#dc3545', '#6f42c1', '#fd7e14', '#20c997', '#e83e8c']
    })
    
    # Prepare DPM chart data for Man Hours
    mh_dpm_breakdown = man_hours.get('dpm_breakdown', [])
    mh_dpm_labels = [item['dpm_name'] for item in mh_dpm_breakdown]
    mh_dpm_hours = [item['hours'] for item in mh_dpm_breakdown]
    
    mh_dpm_chart_json = json.dumps({
        'labels': mh_dpm_labels,
        'data': mh_dpm_hours,
        'colors': ['#28a745', '#17a2b8', '#ffc107', '#dc3545', '#6f42c1', '#fd7e14']
    })

    context = {
        'report_data': report_data,
        'sales_confirmed': sales_confirmed,
        'first_cut_deliveries': first_cut_deliveries,
        'final_deliveries': final_deliveries,
        'man_hours': man_hours,
        'filters': filters,
        'products': products,
        'dpms': dpms,
        'cities': cities,
        'product_chart_json': product_chart_json,
        'dpm_chart_json': dmp_chart_json,
        'fcd_product_chart_json': fcd_product_chart_json,
        'fcd_dpm_chart_json': fcd_dpm_chart_json,
        'fd_product_chart_json': fd_product_chart_json,
        'fd_dpm_chart_json': fd_dpm_chart_json,
        'mh_product_chart_json': mh_product_chart_json,
        'mh_dpm_chart_json': mh_dpm_chart_json,
        'title': 'General Business Report'
    }
    
    return render(request, 'projects/reports/general_report.html', context)


@login_required
def tat_analytics_simple(request):
    """
    Simplified TAT Analytics Dashboard for Senior Managers.
    Focuses on TAT adherence percentages across different dimensions.
    """
    # Check permissions - allow SENIOR_MANAGER, DPM, and VIDEO_PM
    redirect_response = ensure_has_management_access(request)
    if redirect_response:
        return redirect_response
    
    # Get filter data for dropdowns
    from .models import Product
    products = Product.objects.all().order_by('name')
    dpms = User.objects.filter(role='DPM').order_by('first_name', 'last_name')
    
    # Build filters dictionary
    filters = {}
    applied_filters = {}
    
    # Date filters
    if request.GET.get('date_from'):
        try:
            filters['date_from'] = datetime.strptime(request.GET.get('date_from'), '%Y-%m-%d').date()
            applied_filters['date_from'] = filters['date_from']
        except ValueError:
            pass
    
    if request.GET.get('date_to'):
        try:
            filters['date_to'] = datetime.strptime(request.GET.get('date_to'), '%Y-%m-%d').date()
            applied_filters['date_to'] = filters['date_to']
        except ValueError:
            pass
    
    # Product filter
    if request.GET.get('product'):
        try:
            product_obj = Product.objects.get(id=request.GET.get('product'))
            filters['product'] = product_obj.name
            applied_filters['product'] = request.GET.get('product')
        except (Product.DoesNotExist, ValueError):
            pass
    
    # DPM filter
    if request.GET.get('dpm'):
        try:
            dpm_obj = User.objects.get(id=request.GET.get('dpm'), role='DPM')
            filters['dpm'] = dpm_obj.username
            applied_filters['dpm'] = request.GET.get('dpm')
        except (User.DoesNotExist, ValueError):
            pass
    
    # Add default date range if no filters provided to improve performance
    if not any([request.GET.get('date_from'), request.GET.get('date_to'), 
                request.GET.get('product'), request.GET.get('dpm')]):
        # Default to July 1, 2025
        from datetime import date
        default_start = date(2025, 7, 1)
        filters['date_from'] = default_start
        applied_filters['date_from'] = default_start
    
    # Get simplified dashboard data
    dashboard_data = TATAnalyticsService.get_simplified_tat_dashboard(filters)
    
    if not dashboard_data['success']:
        messages.error(request, "Error loading TAT analytics")
        dashboard_data = {
            'adherence_data': {
                'summary': {},
                'dpm_wise': [],
                'city_wise': [],
                'product_wise': []
            }
        }
    
    import json
    
    # JSON encode trend data for safe JavaScript rendering
    trend_data = dashboard_data.get('trend_data', {})
    trend_data_json = {
        'labels': json.dumps(trend_data.get('labels', [])),
        'data': json.dumps(trend_data.get('data', []))
    }
    
    # Ensure adherence data has proper defaults for JavaScript rendering
    adherence_data = dashboard_data.get('adherence_data', {})
    summary = adherence_data.get('summary', {})
    
    # Convert to integers to ensure proper JavaScript rendering
    within_tat = int(summary.get('within_tat', 0) or 0)
    beyond_tat = int(summary.get('beyond_tat', 0) or 0)
    total_projects = int(summary.get('total', 0) or 0)  # Use 'total', not 'total_projects'
    adherence_percentage = float(summary.get('adherence_percentage', 0) or 0)
    
    adherence_summary = {
        'within_tat': within_tat,
        'beyond_tat': beyond_tat,
        'total': total_projects,
        'adherence_percentage': adherence_percentage
    }
    
    # JSON encode the TAT distribution data for safe JavaScript embedding
    tat_distribution_json = {
        'within_tat': json.dumps(within_tat),
        'beyond_tat': json.dumps(beyond_tat)
    }
    
    # Pipeline vs Delivered adherence data with JSON encoding
    pipeline_delivered_data = adherence_data.get('pipeline_delivered', {})
    pipeline_data = pipeline_delivered_data.get('pipeline', {})
    delivered_data = pipeline_delivered_data.get('delivered', {})
    
    pipeline_json = {
        'within_tat': json.dumps(int(pipeline_data.get('within_tat', 0) or 0)),
        'beyond_tat': json.dumps(int(pipeline_data.get('beyond_tat', 0) or 0)),
        'total_projects': int(pipeline_data.get('total_projects', 0) or 0),
        'adherence_percentage': float(pipeline_data.get('adherence_percentage', 0) or 0)
    }
    
    delivered_json = {
        'within_tat': json.dumps(int(delivered_data.get('within_tat', 0) or 0)),
        'beyond_tat': json.dumps(int(delivered_data.get('beyond_tat', 0) or 0)),
        'total_projects': int(delivered_data.get('total_projects', 0) or 0),
        'adherence_percentage': float(delivered_data.get('adherence_percentage', 0) or 0)
    }
    
    context = {
        'adherence_data': {
            'summary': adherence_summary,
            'dpm_wise': adherence_data.get('dpm_wise', []),
            'city_wise': adherence_data.get('city_wise', []),
            'region_wise': adherence_data.get('region_wise', []),
            'product_wise': adherence_data.get('product_wise', []),
            'pipeline_delivered': {
                'pipeline': pipeline_json,
                'delivered': delivered_json
            }
        },
        'trend_data': trend_data,
        'trend_data_json': trend_data_json,
        'tat_distribution_json': tat_distribution_json,
        'total_projects': dashboard_data.get('total_projects', 0),
        'filters': filters,
        'applied_filters': applied_filters,
        'products': products,
        'dpms': dpms,
        'title': 'TAT Adherence Dashboard'
    }
    
    return render(request, 'projects/reports/tat_adherence_simple.html', context)


@login_required
def ageing_report_dashboard(request):
    """
    Project Ageing Report Dashboard for Senior Managers.
    Shows distribution of projects by age categories with quantity weighting.
    """
    # Check permissions - allow SENIOR_MANAGER, DPM, and VIDEO_PM
    redirect_response = ensure_has_management_access(request)
    if redirect_response:
        return redirect_response
    
    # Parse filters from request
    filters = {}
    if request.GET.get('date_from'):
        try:
            filters['date_from'] = datetime.strptime(request.GET.get('date_from'), '%Y-%m-%d').date()
        except ValueError:
            pass
    else:
        # Default to July 1, 2025 if no date_from provided
        filters['date_from'] = date(2025, 7, 1)
    
    if request.GET.get('date_to'):
        try:
            filters['date_to'] = datetime.strptime(request.GET.get('date_to'), '%Y-%m-%d').date()
        except ValueError:
            pass
    
    if request.GET.get('product'):
        filters['product'] = request.GET.get('product')
    
    if request.GET.get('dpm'):
        filters['dpm'] = request.GET.get('dpm')
    
    if request.GET.get('city'):
        filters['city'] = request.GET.get('city')
    
    # Get ageing data from service
    from .services import AgeingReportService
    dashboard_data = AgeingReportService.get_ageing_dashboard(filters)
    
    if not dashboard_data.get('success'):
        messages.error(request, f"Error generating ageing report: {dashboard_data.get('error', 'Unknown error')}")
        dashboard_data = {
            'category_totals': {'1+ year': 0, '6+ months': 0, '3+ months': 0, 'less than 3 months': 0},
            'dpm_breakdown': [],
            'product_breakdown': [],
            'total_projects': 0,
            'project_count': 0
        }
    
    # Prepare chart data for frontend
    category_labels = ['1+ year', '6+ months', '3+ months', 'less than 3 months']
    category_totals = dashboard_data.get('category_totals', {})
    category_data = [
        category_totals.get('one_plus_year', 0),
        category_totals.get('six_plus_months', 0), 
        category_totals.get('three_plus_months', 0),
        category_totals.get('less_than_3_months', 0)
    ]
    category_colors = ['#dc3545', '#fd7e14', '#0dcaf0', '#198754']  # Red, Orange, Blue, Green
    
    category_chart_json = json.dumps({
        'labels': category_labels,
        'data': category_data,
        'colors': category_colors
    })
    
    # Get filter options for dropdowns
    products = Product.objects.all().values_list('name', flat=True).distinct()
    dpms = User.objects.filter(role='DPM').values_list('username', flat=True)
    cities = City.objects.all().values_list('name', flat=True).distinct()
    
    # Prepare DPM chart data arrays
    dpm_breakdown = dashboard_data.get('dpm_breakdown', [])
    dpm_labels = [dpm.get('dpm_name', '') for dpm in dpm_breakdown]
    dpm_one_year = [dpm.get('one_plus_year', 0) for dpm in dpm_breakdown]
    dpm_six_months = [dpm.get('six_plus_months', 0) for dpm in dpm_breakdown]
    dpm_three_months = [dpm.get('three_plus_months', 0) for dpm in dpm_breakdown]
    dpm_less_three = [dpm.get('less_than_3_months', 0) for dpm in dpm_breakdown]
    
    context = {
        'ageing_data': dashboard_data,
        'category_chart_json': category_chart_json,
        'total_projects': dashboard_data.get('total_projects', 0),
        'project_count': dashboard_data.get('project_count', 0),
        'filters': filters,
        'products': products,
        'dpms': dpms,
        'cities': cities,
        'dpm_labels': json.dumps(dpm_labels),
        'dpm_one_year': json.dumps(dpm_one_year),
        'dmp_six_months': json.dumps(dpm_six_months),
        'dpm_three_months': json.dumps(dpm_three_months),
        'dpm_less_three': json.dumps(dpm_less_three),
        'title': 'Project Ageing Report'
    }
    
    return render(request, 'projects/reports/ageing_report.html', context)


@login_required
def tat_analytics_dashboard(request):
    """
    Legacy TAT Analytics Dashboard - redirects to simplified version.
    """
    # Redirect to the simplified dashboard
    return redirect('projects:tat_analytics_simple')


def export_tat_data(request, filters, format='csv'):
    """
    Export TAT analytics data in CSV or Excel format.
    """
    from accounts.services import ensure_has_management_access
    from django.http import FileResponse
    import os
    
    try:
        ensure_has_management_access(request.user)
    except PermissionError as e:
        messages.error(request, str(e))
        return redirect('projects:tat_analytics_dashboard')
    
    # Get projects with filters applied
    projects_queryset = Project.objects.select_related(
        'product', 'current_status', 'dpm', 'city', 'project_incharge'
    ).prefetch_related('status_history__status')
    
    # Apply same filters as dashboard
    if filters:
        if filters.get('date_from'):
            projects_queryset = projects_queryset.filter(purchase_date__gte=filters['date_from'])
        if filters.get('date_to'):
            projects_queryset = projects_queryset.filter(purchase_date__lte=filters['date_to'])
        if filters.get('product'):
            projects_queryset = projects_queryset.filter(product_id=filters['product'])
        if filters.get('dpm'):
            projects_queryset = projects_queryset.filter(dpm_id=filters['dpm'])
        if filters.get('city'):
            projects_queryset = projects_queryset.filter(city_id=filters['city'])
        if filters.get('project_type'):
            if filters['project_type'] == 'delivered':
                projects_queryset = projects_queryset.filter(
                    current_status__category_two__iexact='Final Delivery'
                )
            elif filters['project_type'] == 'pipeline':
                projects_queryset = projects_queryset.exclude(
                    current_status__category_two__iexact='Final Delivery'
                )
    
    # Get projects with TAT data
    projects_with_tat = ProjectService.get_projects_with_tat_data(projects_queryset)
    
    # Export data
    success, result = TATAnalyticsService.export_tat_data(projects_with_tat, format)
    
    if success:
        # Return file download
        response = FileResponse(
            open(result, 'rb'),
            as_attachment=True,
            filename=os.path.basename(result)
        )
        return response
    else:
        messages.error(request, f"Export failed: {result}")
        return redirect('projects:tat_analytics_dashboard')


@login_required
def export_status_history(request):
    """
    Export Project Status History data with enriched project information.
    Supports both CSV and Excel formats with filtering options.
    """
    # Check permissions
    ensure_has_management_access(request)
    
    # Get export format (default to excel)
    export_format = request.GET.get('format', 'excel').lower()
    
    # Get filters from request
    filters = {}
    if request.GET.get('start_date'):
        try:
            filters['start_date'] = datetime.strptime(request.GET.get('start_date'), '%Y-%m-%d').date()
        except ValueError:
            pass
    
    if request.GET.get('end_date'):
        try:
            filters['end_date'] = datetime.strptime(request.GET.get('end_date'), '%Y-%m-%d').date()
        except ValueError:
            pass
    
    if request.GET.get('product'):
        filters['product'] = request.GET.get('product')
    
    if request.GET.get('dpm'):
        filters['dpm'] = request.GET.get('dpm')
    
    if request.GET.get('city'):
        filters['city'] = request.GET.get('city')
    
    if request.GET.get('project_type'):
        filters['project_type'] = request.GET.get('project_type')
    
    # Get enriched status history data
    status_history_queryset = ProjectService.get_enriched_status_history_data(filters)
    export_data = ProjectService.prepare_status_history_export_data(status_history_queryset)
    
    if not export_data:
        messages.warning(request, "No status history data found for the selected filters.")
        return redirect('projects:reports_dashboard')
    
    # Create filename with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    if export_format == 'csv':
        return _export_status_history_csv(export_data, timestamp)
    else:
        return _export_status_history_excel(export_data, timestamp)


def _export_status_history_csv(export_data, timestamp):
    """Export status history data as CSV."""
    import csv
    from io import StringIO
    
    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    filename = f"project_status_history_{timestamp}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Get field names from first record
    if export_data:
        fieldnames = list(export_data[0].keys())
        
        writer = csv.DictWriter(response, fieldnames=fieldnames)
        writer.writeheader()
        
        for row in export_data:
            writer.writerow(row)
    
    return response


def _export_status_history_excel(export_data, timestamp):
    """Export status history data as Excel with formatting."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Project Status History"
    
    if not export_data:
        # Empty workbook
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheet.sheet'
        )
        filename = f"project_status_history_{timestamp}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )
    
    # Get headers from first record
    headers = list(export_data[0].keys())
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header.replace('_', ' ').title())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
        
        # Set column width based on header length
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = max(15, len(str(header)) + 2)
    
    # Write data
    for row_idx, record in enumerate(export_data, 2):
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=record[header])
            cell.border = border
            
            # Center align numeric and date columns
            if header in ['quantity', 'expected_tat_days', 'actual_tat_days'] or 'date' in header:
                cell.alignment = center_alignment
    
    # Add title and summary info
    ws.insert_rows(1, 3)
    
    # Title
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = "Project Status History Export"
    title_cell.font = Font(bold=True, size=16, color="4472C4")
    title_cell.alignment = center_alignment
    
    # Export info
    ws.merge_cells('A2:E2')
    info_cell = ws['A2']
    info_cell.value = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Total Records: {len(export_data)}"
    info_cell.font = Font(italic=True)
    info_cell.alignment = center_alignment
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheet.sheet'
    )
    filename = f"project_status_history_{timestamp}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response