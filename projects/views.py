#projects/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.urls import reverse
from django.http import JsonResponse, Http404
from urllib.parse import quote
from .forms import (
    ProjectStatusUpdateForm, ProjectFilterForm, ProjectCreateForm, ProjectTaskForm, 
    TaskAssignmentForm, TaskAssignmentUpdateForm, ProjectManagementForm, 
    AddMiscHoursForm, EditMiscHoursForm, TimerStopForm, ManualTimeEntryForm, 
    EditSessionDurationForm, DailyRosterFilterForm, TaskAssignmentFilterForm,
    DeliveredProjectFilterForm
)
from .services import ProjectService, TATAnalyticsService
from accounts.models import User
from locations.models import Region, City
from django.http import JsonResponse
from .models import (
    Project, ProjectStatusOption, ProjectTask, TaskAssignment, 
    ProjectStatusHistory, ActiveTimer, TimeSession, DailyTimeTotal, 
    ProjectDelivery, Product, MiscHours
)
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.db.models import Subquery, OuterRef, F, Avg, Count, Q, Sum
from django.core.exceptions import ValidationError
import uuid
from datetime import date, timedelta, datetime
from django.utils import timezone
import json
# Export functionality imports
import csv
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter




def ensure_is_dpm(request, project):
    """
    Verify that the current user is a DPM.

    Args:
        request: The HTTP request
        project: The project object (used for redirect)

    Returns:
        Response or None: Redirect response if check fails, None if successful
    """
    if request.user.role != 'DPM':
        messages.error(request, "Only a DPM can perform this action.")
        return redirect('projects:project_detail', project_id=project.id)
    return None

def ensure_has_management_access(request):
    """
    Verify that the current user has management access (DPM, VIDEO_PM, or SENIOR_MANAGER).
    Senior Managers have read-only access to reporting features.

    Args:
        request: The HTTP request

    Returns:
        Response or None: Redirect response if check fails, None if successful
    """
    if request.user.role not in ['DPM', 'VIDEO_PM', 'SENIOR_MANAGER']:
        messages.error(request, "You don't have permission to access this page.")
        return redirect('projects:project_list')
    return None

def ensure_has_full_management_access(request):
    """
    Verify that the current user has full management access (DPM or VIDEO_PM only).
    Senior Managers are excluded from this check as they have read-only access.

    Args:
        request: The HTTP request

    Returns:
        Response or None: Redirect response if check fails, None if successful
    """
    if request.user.role not in ['DPM', 'VIDEO_PM']:
        messages.error(request, "You don't have permission to perform this action.")
        return redirect('projects:project_list')
    return None

@login_required
def create_project(request):
    """
    View for creating a new project. Handles form processing and
    delegates business logic to the service layer.
    Only accessible by DPMs and Senior Managers.
    """
    # Check if user is a DPM
    if request.user.role != 'DPM':
        messages.error(request, "Access denied. Only Project Managers can create projects.")
        return redirect('home')
    
    if request.method == 'POST':
        # Create and validate form
        form = ProjectCreateForm(request.POST, user=request.user)

        if form.is_valid():
            try:
                # Form's save method delegates to service
                project = form.save()
                messages.success(request, f'Project "{project.project_name}" has been created successfully.')
                return redirect('projects:project_detail', project_id=project.id)
            except ValidationError as e:
                # Handle service-layer validation errors
                messages.error(request, f"Error creating project: {e}")
        else:
            # Form validation failed
            messages.error(request, "Please correct the errors below.")
    else:
        # For GET requests, create an empty form
        form = ProjectCreateForm(user=request.user)

    # Render the template with the form
    return render(request, 'projects/create_project.html', {'form': form, 'title': 'Create New Project'})


@login_required
def project_detail(request, project_id):
    """Display detailed information about a specific project. Only accessible by DPMs and Senior Managers."""
    # Check if user is a DPM or Senior Manager
    if request.user.role not in ['DPM', 'SENIOR_MANAGER']:
        messages.error(request, "Access denied. Only Project Managers and Senior Managers can view project details.")
        return redirect('home')
    
    # Get project data with TAT calculation using service
    try:
        project, tat_data = ProjectService.get_project_with_tat_data(project_id)
        
        # Get status history in chronological order (oldest first)
        status_history = project.status_history.all().order_by('changed_at')
        
    except Http404:
        messages.error(request, "Project not found.")
        return redirect('projects:project_list')
    except Exception as e:
        messages.error(request, f"Error loading project: {str(e)}")
        return redirect('projects:project_list')

    # Get status options for the modal
    status_options = ProjectStatusOption.objects.filter(is_active=True).order_by('order')

    # Prepare the form for the modal
    form = ProjectStatusUpdateForm(initial={'status': project.current_status})

    # Prepare context
    context = {
        'project': project,
        'tat_data': tat_data,
        'status_history': status_history,
        'status_options': status_options,
        'form': form,
        'today': timezone.now().date(),
        'title': f'Project: {project.project_name}'
    }

    return render(request, 'projects/project_detail.html', context)

@login_required
def update_project_status(request, project_id):
    """
    Handle status updates for a project.
    Now supports both regular and AJAX requests.
    Only accessible by DPMs and Senior Managers.
    """
    # Check if user is a DPM first
    if request.user.role != 'DPM':
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': 'Access denied. Only Project Managers can update project status.'}, status=403)
        messages.error(request, "Access denied. Only Project Managers can update project status.")
        return redirect('home')
    
    # Get the project first
    success, result = ProjectService.get_project(project_id)

    if not success:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': result}, status=404)
        messages.error(request, result)
        return redirect('projects:project_list')

    project = result

    # Check if it's an AJAX request to get status options
    if request.method == 'GET' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        # Return available status options as JSON
        statuses = ProjectStatusOption.objects.filter(is_active=True).order_by('order')
        status_options = [{'id': str(s.id), 'name': s.name} for s in statuses]
        return JsonResponse({'status_options': status_options})

    if request.method == 'POST':
        form = ProjectStatusUpdateForm(request.POST)
        if form.is_valid():
            status_id = form.cleaned_data['status'].id
            comments = form.cleaned_data['comments']
            status_date = form.cleaned_data['status_date']

            success, result = ProjectService.update_project_status(
                project_id=project_id,
                status_id=status_id,
                user=request.user,
                comments=comments,
                status_date=status_date
            )

            if success:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True,
                        'message': 'Project status updated successfully.',
                        'new_status': result.current_status.name
                    })
                messages.success(request, "Project status updated successfully.")
                return redirect('projects:project_detail', project_id=project_id)
            else:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'message': result}, status=400)
                messages.error(request, result)
        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                # Return form errors as JSON
                errors = {}
                for field, field_errors in form.errors.items():
                    errors[field] = list(field_errors)
                return JsonResponse({'success': False, 'errors': errors}, status=400)
            messages.error(request, "Please correct the form errors.")
    else:
        # Pre-select current status in form
        form = ProjectStatusUpdateForm(initial={'status': project.current_status})

    # For non-AJAX requests, render the original template
    return render(request, 'projects/status_update.html', {
        'form': form,
        'project': project,
        'today': timezone.now().date(),
        'title': f'Update Status: {project.project_name}'
    })


@login_required
def project_list(request):
    """
    Displays a filterable list of pipeline projects (not yet delivered).
    Only accessible by DPMs and Senior Managers.
    """
    # Check if user is a DPM or Senior Manager
    if request.user.role not in ['DPM', 'SENIOR_MANAGER']:
        messages.error(request, "Access denied. This page is only for Project Managers and Senior Managers.")
        return redirect('home')
    
    # Get filter parameters from request
    search_query = request.GET.get('search', '')
    status = request.GET.get('status', '')
    product = request.GET.get('product', '')
    region = request.GET.get('region', '')
    city = request.GET.get('city', '')
    dpm = request.GET.get('dpm', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    page = request.GET.get('page', 1)

    # Convert date strings to date objects if provided
    date_from_obj = None
    date_to_obj = None
    if date_from:
        try:
            from datetime import datetime
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
        except ValueError:
            pass
    if date_to:
        try:
            from datetime import datetime
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
        except ValueError:
            pass

    # Get pipeline projects using service (exclude statuses with category_two = 'Final Delivery')
    success, result = ProjectService.get_project_list(
        search_query=search_query,
        status=status,
        product=product,
        region=region,
        city=city,
        dpm=dpm,
        date_from=date_from_obj,
        date_to=date_to_obj,
        page=page,
        project_type='pipeline'  # Only get pipeline projects
    )

    if not success:
        messages.error(request, result)
        return redirect('home')

    projects, filters_applied = result

    # Get filter options
    success, filter_options_result = ProjectService.get_filter_options()
    if not success:
        messages.warning(request, filter_options_result)
        filter_options = {
            'statuses': [],
            'products': [],
            'cities': [],
            'regions': [],
            'dpms': []
        }
    else:
        filter_options = filter_options_result

    # Create filter form with current values
    filter_form = ProjectFilterForm(initial=filters_applied)

    # Update city queryset based on selected region
    if region:
        filter_form.fields['city'].queryset = City.objects.filter(region_id=region)

    # Get display names for applied filters for the template
    filters_applied_display = {}
    if filters_applied.get('status'):
        try:
            filters_applied_display['status'] = ProjectStatusOption.objects.get(id=filters_applied['status']).name
        except ProjectStatusOption.DoesNotExist:
            pass
    if filters_applied.get('product'):
        try:
            filters_applied_display['product'] = Product.objects.get(id=filters_applied['product']).name
        except Product.DoesNotExist:
            pass
    if filters_applied.get('region'):
        try:
            filters_applied_display['region'] = Region.objects.get(id=filters_applied['region']).name
        except Region.DoesNotExist:
            pass
    if filters_applied.get('city'):
        try:
            filters_applied_display['city'] = City.objects.get(id=filters_applied['city']).name
        except City.DoesNotExist:
            pass
    if filters_applied.get('dpm'):
        try:
            dpm_user = User.objects.get(id=filters_applied['dpm'])
            filters_applied_display['dpm'] = dpm_user.get_full_name() or dpm_user.username
        except User.DoesNotExist:
            pass

    context = {
        'projects': projects,
        'filter_form': filter_form,
        'filters_applied': filters_applied,
        'filters_applied_display': filters_applied_display,
        'filter_options': filter_options,
        'title': 'Pipeline Projects',
        'is_pipeline': True  # Add flag to identify page type
    }

    return render(request, 'projects/project_list.html', context)


@login_required
def all_projects(request):
    """
    Displays a filterable list of ALL projects (both pipeline and delivered).
    Only accessible by DPMs and Senior Managers.
    """
    # Check if user is a DPM or Senior Manager
    if request.user.role not in ['DPM', 'SENIOR_MANAGER']:
        messages.error(request, "Access denied. This page is only for Project Managers and Senior Managers.")
        return redirect('home')
    
    # Get filter parameters from request
    search_query = request.GET.get('search', '')
    status = request.GET.get('status', '')
    product = request.GET.get('product', '')
    region = request.GET.get('region', '')
    city = request.GET.get('city', '')
    dpm = request.GET.get('dpm', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    type_filter = request.GET.get('type', 'all')
    page = request.GET.get('page', 1)

    # Convert date strings to date objects if provided
    date_from_obj = None
    date_to_obj = None
    if date_from:
        try:
            from datetime import datetime
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
        except ValueError:
            pass
    if date_to:
        try:
            from datetime import datetime
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
        except ValueError:
            pass

    # Determine project_type based on type_filter
    if type_filter == 'pipeline':
        project_type = 'pipeline'
    elif type_filter == 'delivered':
        project_type = 'delivered'
    else:
        project_type = 'all'
    
    # Get projects using service with type filtering
    success, result = ProjectService.get_project_list(
        search_query=search_query,
        status=status,
        product=product,
        region=region,
        city=city,
        dpm=dpm,
        date_from=date_from_obj,
        date_to=date_to_obj,
        page=page,
        project_type=project_type
    )

    if not success:
        messages.error(request, result)
        return redirect('home')

    projects, filters_applied = result

    # Get total count of all projects (unfiltered)
    total_success, total_result = ProjectService.get_project_list(
        project_type='all',
        page=1,
        items_per_page=1  # We only need the count
    )
    total_count = 0
    if total_success:
        total_count = total_result[0].paginator.count

    # Get filter options
    success, filter_options_result = ProjectService.get_filter_options()
    if not success:
        messages.warning(request, filter_options_result)
        filter_options = {
            'statuses': [],
            'products': [],
            'cities': [],
            'regions': [],
            'dpms': []
        }
    else:
        filter_options = filter_options_result

    # Build filter text for display
    filter_parts = []
    if search_query:
        filter_parts.append(f"Search: {search_query}")
    if status:
        status_name = next((s.name for s in filter_options['statuses'] if str(s.id) == status), status)
        filter_parts.append(f"Status: {status_name}")
    if product:
        product_name = next((p.name for p in filter_options['products'] if str(p.id) == product), product)
        filter_parts.append(f"Product: {product_name}")
    if region:
        region_name = next((r.name for r in filter_options['regions'] if str(r.id) == region), region)
        filter_parts.append(f"Region: {region_name}")
    if city:
        city_name = next((c.name for c in filter_options['cities'] if str(c.id) == city), city)
        filter_parts.append(f"City: {city_name}")
    if dpm:
        dpm_name = next((d.username for d in filter_options['dpms'] if str(d.id) == dpm), dpm)
        filter_parts.append(f"DPM: {dpm_name}")
    if type_filter and type_filter != 'all':
        type_display = 'Pipeline' if type_filter == 'pipeline' else 'Delivered'
        filter_parts.append(f"Type: {type_display}")
    if date_from:
        filter_parts.append(f"From: {date_from}")
    if date_to:
        filter_parts.append(f"To: {date_to}")

    filter_text = " | ".join(filter_parts) if filter_parts else "No filters applied"
    
    # Check if any date filters are active
    date_filter_active = bool(date_from or date_to)

    context = {
        'projects': projects,
        'filter_options': filter_options,
        'current_filters': {
            'search': search_query,
            'status': status,
            'product': product,
            'region': region,
            'city': city,
            'dpm': dpm,
            'type': type_filter,
            'date_from': date_from,
            'date_to': date_to,
        },
        'filters_applied': filters_applied,
        'filter_text': filter_text,
        'date_filter_active': date_filter_active,
        'total_count': total_count,
        'filtered_count': projects.paginator.count
    }

    return render(request, 'projects/all_projects.html', context)


@login_required
def export_all_projects(request):
    """
    Export all projects (both pipeline and delivered) to CSV or XLSX format.
    Uses the same filtering logic as all_projects view.
    Only accessible by DPMs and Senior Managers.
    """
    # Check if user is a DPM or Senior Manager
    if request.user.role not in ['DPM', 'SENIOR_MANAGER']:
        messages.error(request, "Access denied. This page is only for Project Managers and Senior Managers.")
        return redirect('home')
    
    # Get export format from request (default to CSV)
    export_format = request.GET.get('format', 'csv').lower()
    
    # Get filter parameters from request (same as all_projects view)
    search_query = request.GET.get('search', '')
    status = request.GET.get('status', '')
    product = request.GET.get('product', '')
    region = request.GET.get('region', '')
    city = request.GET.get('city', '')
    dpm = request.GET.get('dpm', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    # Convert date strings to date objects if provided
    date_from_obj = None
    date_to_obj = None
    if date_from:
        try:
            from datetime import datetime
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
        except ValueError:
            pass
    if date_to:
        try:
            from datetime import datetime
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
        except ValueError:
            pass

    # Get ALL projects (no pagination for export)
    success, result = ProjectService.get_project_list(
        search_query=search_query,
        status=status,
        product=product,
        region=region,
        city=city,
        dpm=dpm,
        date_from=date_from_obj,
        date_to=date_to_obj,
        page=1,
        items_per_page=10000,  # Large number to get all results
        project_type='all'  # Get all projects (both pipeline and delivered)
    )

    if not success:
        messages.error(request, f"Error exporting projects: {result}")
        return redirect('projects:all_projects')

    projects, _ = result
    projects_list = projects.object_list if hasattr(projects, 'object_list') else projects

    # Define the columns for export
    headers = [
        'HS ID', 'Opportunity ID', 'Project Name', 'Builder Name', 'City', 'Region',
        'Product', 'Product Subcategory', 'Package ID', 'Quantity', 'Purchase Date',
        'Sales Confirmation Date', 'Expected TAT', 'Account Manager', 'DPM',
        'Current Status', 'Project Incharge', 'Expected Completion Date',
        'Delivery Performance Rating', 'Created At', 'Project Type'
    ]

    if export_format == 'xlsx':
        return _export_all_to_xlsx(projects_list, headers)
    else:
        return _export_all_to_csv(projects_list, headers)


def _export_all_to_csv(projects, headers):
    """Helper function to export all projects to CSV format."""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="all_projects_{date.today().strftime("%Y%m%d")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(headers)
    
    for project in projects:
        row = [
            project.hs_id,
            project.opportunity_id,
            project.project_name,
            project.builder_name,
            project.city.name if project.city else '',
            project.city.region.name if project.city and project.city.region else '',
            project.product.name if project.product else '',
            project.product_subcategory.name if project.product_subcategory else '',
            project.package_id or '',
            project.quantity,
            project.purchase_date.strftime('%Y-%m-%d') if project.purchase_date else '',
            project.sales_confirmation_date.strftime('%Y-%m-%d') if project.sales_confirmation_date else '',
            project.expected_tat,
            project.account_manager,
            project.dpm.get_full_name() if project.dpm else '',
            project.current_status.name if project.current_status else '',
            project.project_incharge.get_full_name() if project.project_incharge else '',
            project.expected_completion_date.strftime('%Y-%m-%d') if project.expected_completion_date else '',
            str(project.delivery_performance_rating) if project.delivery_performance_rating else '',
            project.created_at.strftime('%Y-%m-%d %H:%M:%S') if project.created_at else '',
            'Delivered' if project.is_delivered else 'Pipeline'
        ]
        writer.writerow(row)
    
    return response


def _export_all_to_xlsx(projects, headers):
    """Helper function to export all projects to XLSX format."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "All Projects"
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Write data
    for row_num, project in enumerate(projects, 2):
        data = [
            project.hs_id,
            project.opportunity_id,
            project.project_name,
            project.builder_name,
            project.city.name if project.city else '',
            project.city.region.name if project.city and project.city.region else '',
            project.product.name if project.product else '',
            project.product_subcategory.name if project.product_subcategory else '',
            project.package_id or '',
            project.quantity,
            project.purchase_date.strftime('%Y-%m-%d') if project.purchase_date else '',
            project.sales_confirmation_date.strftime('%Y-%m-%d') if project.sales_confirmation_date else '',
            project.expected_tat,
            project.account_manager,
            project.dpm.get_full_name() if project.dpm else '',
            project.current_status.name if project.current_status else '',
            project.project_incharge.get_full_name() if project.project_incharge else '',
            project.expected_completion_date.strftime('%Y-%m-%d') if project.expected_completion_date else '',
            str(project.delivery_performance_rating) if project.delivery_performance_rating else '',
            project.created_at.strftime('%Y-%m-%d %H:%M:%S') if project.created_at else '',
            'Delivered' if project.is_delivered else 'Pipeline'
        ]
        
        for col_num, value in enumerate(data, 1):
            worksheet.cell(row=row_num, column=col_num, value=value)
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="all_projects_{date.today().strftime("%Y%m%d")}.xlsx"'
    
    workbook.save(response)
    return response


@login_required
def export_pipeline_projects(request):
    """
    Export pipeline projects (non-delivered) to CSV or XLSX format.
    Uses the same filtering logic as project_list view.
    Only accessible by DPMs and Senior Managers.
    """
    # Check if user is a DPM or Senior Manager
    if request.user.role not in ['DPM', 'SENIOR_MANAGER']:
        messages.error(request, "Access denied. This page is only for Project Managers and Senior Managers.")
        return redirect('home')
    
    # Get export format from request (default to CSV)
    export_format = request.GET.get('format', 'csv').lower()
    
    # Get filter parameters from request (same as project_list view)
    search_query = request.GET.get('search', '')
    status = request.GET.get('status', '')
    product = request.GET.get('product', '')
    region = request.GET.get('region', '')
    city = request.GET.get('city', '')
    dpm = request.GET.get('dpm', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    # Convert date strings to date objects if provided
    date_from_obj = None
    date_to_obj = None
    if date_from:
        try:
            from datetime import datetime
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
        except ValueError:
            pass
    if date_to:
        try:
            from datetime import datetime
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
        except ValueError:
            pass

    # Get ALL pipeline projects (no pagination for export)
    success, result = ProjectService.get_project_list(
        search_query=search_query,
        status=status,
        product=product,
        region=region,
        city=city,
        dpm=dpm,
        date_from=date_from_obj,
        date_to=date_to_obj,
        page=1,
        items_per_page=10000,  # Large number to get all results
        project_type='pipeline'  # Only get pipeline projects
    )

    if not success:
        messages.error(request, f"Error exporting projects: {result}")
        return redirect('projects:project_list')

    projects, _ = result
    projects_list = projects.object_list if hasattr(projects, 'object_list') else projects

    # Define the columns for export
    headers = [
        'HS ID', 'Opportunity ID', 'Project Name', 'Builder Name', 'City', 'Region',
        'Product', 'Product Subcategory', 'Package ID', 'Quantity', 'Purchase Date',
        'Sales Confirmation Date', 'Expected TAT', 'Account Manager', 'DPM',
        'Current Status', 'Project Incharge', 'Expected Completion Date',
        'Delivery Performance Rating', 'Created At'
    ]

    if export_format == 'xlsx':
        return _export_to_xlsx(projects_list, headers)
    else:
        return _export_to_csv(projects_list, headers)


def _export_to_csv(projects, headers):
    """Helper function to export projects to CSV format."""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="pipeline_projects_{date.today().strftime("%Y%m%d")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(headers)
    
    for project in projects:
        row = [
            project.hs_id,
            project.opportunity_id,
            project.project_name,
            project.builder_name,
            project.city.name if project.city else '',
            project.city.region.name if project.city and project.city.region else '',
            project.product.name if project.product else '',
            project.product_subcategory.name if project.product_subcategory else '',
            project.package_id or '',
            project.quantity,
            project.purchase_date.strftime('%Y-%m-%d') if project.purchase_date else '',
            project.sales_confirmation_date.strftime('%Y-%m-%d') if project.sales_confirmation_date else '',
            project.expected_tat,
            project.account_manager,
            project.dpm.get_full_name() if project.dpm else '',
            project.current_status.name if project.current_status else '',
            project.project_incharge.get_full_name() if project.project_incharge else '',
            project.expected_completion_date.strftime('%Y-%m-%d') if project.expected_completion_date else '',
            str(project.delivery_performance_rating) if project.delivery_performance_rating else '',
            project.created_at.strftime('%Y-%m-%d %H:%M:%S') if project.created_at else ''
        ]
        writer.writerow(row)
    
    return response


def _export_to_xlsx(projects, headers):
    """Helper function to export projects to XLSX format."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Pipeline Projects"
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Write data
    for row_num, project in enumerate(projects, 2):
        data = [
            project.hs_id,
            project.opportunity_id,
            project.project_name,
            project.builder_name,
            project.city.name if project.city else '',
            project.city.region.name if project.city and project.city.region else '',
            project.product.name if project.product else '',
            project.product_subcategory.name if project.product_subcategory else '',
            project.package_id or '',
            project.quantity,
            project.purchase_date.strftime('%Y-%m-%d') if project.purchase_date else '',
            project.sales_confirmation_date.strftime('%Y-%m-%d') if project.sales_confirmation_date else '',
            project.expected_tat,
            project.account_manager,
            project.dpm.get_full_name() if project.dpm else '',
            project.current_status.name if project.current_status else '',
            project.project_incharge.get_full_name() if project.project_incharge else '',
            project.expected_completion_date.strftime('%Y-%m-%d') if project.expected_completion_date else '',
            str(project.delivery_performance_rating) if project.delivery_performance_rating else '',
            project.created_at.strftime('%Y-%m-%d %H:%M:%S') if project.created_at else ''
        ]
        
        for col_num, value in enumerate(data, 1):
            worksheet.cell(row=row_num, column=col_num, value=value)
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="pipeline_projects_{date.today().strftime("%Y%m%d")}.xlsx"'
    
    workbook.save(response)
    return response


@login_required
def delivered_projects(request):
    """
    Displays a filterable list of delivered projects (category_two = 'Final Delivery').
    Only accessible by DPMs and Senior Managers.
    """
    # Check if user is a DPM or Senior Manager
    if request.user.role not in ['DPM', 'SENIOR_MANAGER']:
        messages.error(request, "Access denied. This page is only for Project Managers and Senior Managers.")
        return redirect('home')
    
    # Get filter parameters from request
    search_query = request.GET.get('search', '')
    status = request.GET.get('status', '')
    product = request.GET.get('product', '')
    region = request.GET.get('region', '')
    city = request.GET.get('city', '')
    dpm = request.GET.get('dpm', '')
    delivery_date_from = request.GET.get('delivery_date_from', '')
    delivery_date_to = request.GET.get('delivery_date_to', '')
    page = request.GET.get('page', 1)

    # Convert date strings to date objects if provided
    date_from_obj = None
    date_to_obj = None
    if delivery_date_from:
        try:
            from datetime import datetime
            date_from_obj = datetime.strptime(delivery_date_from, '%Y-%m-%d').date()
        except ValueError:
            pass
    if delivery_date_to:
        try:
            from datetime import datetime
            date_to_obj = datetime.strptime(delivery_date_to, '%Y-%m-%d').date()
        except ValueError:
            pass

    # Get delivered projects using service
    success, result = ProjectService.get_project_list(
        search_query=search_query,
        status=status,  # Include status filter for delivered projects
        product=product,
        region=region,
        city=city,
        dpm=dpm,
        date_from=date_from_obj,
        date_to=date_to_obj,
        page=page,
        project_type='delivered'  # Only get delivered projects
    )

    if not success:
        messages.error(request, result)
        return redirect('home')

    projects, filters_applied = result

    # Get filter options
    success, filter_options_result = ProjectService.get_filter_options()
    if not success:
        messages.warning(request, filter_options_result)
        filter_options = {
            'statuses': [],
            'products': [],
            'cities': [],
            'regions': [],
            'dpms': []
        }
    else:
        filter_options = filter_options_result

    # Create filter form with current values (map delivery_date fields to the form fields)
    form_initial = {
        'search': filters_applied.get('search'),
        'status': filters_applied.get('status'),
        'product': filters_applied.get('product'),
        'region': filters_applied.get('region'),
        'city': filters_applied.get('city'),
        'dpm': filters_applied.get('dpm'),
        'delivery_date_from': delivery_date_from,
        'delivery_date_to': delivery_date_to,
    }
    filter_form = DeliveredProjectFilterForm(initial=form_initial)

    # Update city queryset based on selected region
    if region:
        filter_form.fields['city'].queryset = City.objects.filter(region_id=region)

    # Get display names for applied filters
    filters_applied_display = {}
    if filters_applied.get('status'):
        try:
            filters_applied_display['status'] = ProjectStatusOption.objects.get(id=filters_applied['status']).name
        except ProjectStatusOption.DoesNotExist:
            pass
    if filters_applied.get('product'):
        try:
            filters_applied_display['product'] = Product.objects.get(id=filters_applied['product']).name
        except Product.DoesNotExist:
            pass
    if filters_applied.get('region'):
        try:
            filters_applied_display['region'] = Region.objects.get(id=filters_applied['region']).name
        except Region.DoesNotExist:
            pass
    if filters_applied.get('city'):
        try:
            filters_applied_display['city'] = City.objects.get(id=filters_applied['city']).name
        except City.DoesNotExist:
            pass
    if filters_applied.get('dpm'):
        try:
            dpm_user = User.objects.get(id=filters_applied['dpm'])
            filters_applied_display['dpm'] = dpm_user.get_full_name() or dpm_user.username
        except User.DoesNotExist:
            pass

    # Update filters_applied to use the delivery date field names for template consistency
    filters_applied_template = {
        'search': filters_applied.get('search'),
        'status': filters_applied.get('status'),
        'product': filters_applied.get('product'),
        'region': filters_applied.get('region'),
        'city': filters_applied.get('city'),
        'dpm': filters_applied.get('dpm'),
        'delivery_date_from': delivery_date_from,
        'delivery_date_to': delivery_date_to,
    }

    context = {
        'projects': projects,
        'filter_form': filter_form,
        'filters_applied': filters_applied_template,
        'filters_applied_display': filters_applied_display,
        'filter_options': filter_options,
        'title': 'Delivered Projects',
        'is_delivered': True  # Add flag to identify page type
    }

    return render(request, 'projects/delivered_projects.html', context)


def get_cities(request):
    """
    API endpoint to get cities for a specific region.
    Used by the dynamic city dropdown in the filter form.
    """
    region_id = request.GET.get('region')
    if region_id:
        cities = City.objects.filter(region_id=region_id).values('id', 'name')
        return JsonResponse(list(cities), safe=False)
    return JsonResponse([], safe=False)


@login_required
def project_management(request, project_id):
    """Display project management page with forms and tasks. Accessible by DPMs and Senior Managers."""
    # Check if user has access (DPM or Senior Manager)
    if request.user.role not in ['DPM', 'SENIOR_MANAGER']:
        messages.error(request, "Access denied. Only Project Managers and Senior Managers can access this page.")
        return redirect('home')
    
    # Get project for permission check
    success, result = ProjectService.get_project(project_id)
    if not success:
        messages.error(request, result)
        return redirect('projects:project_list')

    project = result

    # Get project with tasks using the service method
    success, result = ProjectService.get_project_with_tasks(project_id)

    if not success:
        messages.error(request, result)
        return redirect('projects:project_list')

    # Unpack the result tuple
    project, tasks = result

    # Prepare context
    context = {
        'project': project,
        'tasks': tasks,
        'title': f'Manage Project: {project.project_name}',
        'project_form': ProjectManagementForm(instance=project),
        'task_form': ProjectTaskForm(project=project)
    }

    return render(request, 'projects/project_management.html', context)


@login_required
def create_project_task(request, project_id):
    """Handle task creation only."""
    if request.method != 'POST':
        return redirect('projects:project_management', project_id=project_id)

    # Get project to check configuration
    success, result = ProjectService.get_project(project_id)
    if not success:
        messages.error(request, result)
        return redirect('projects:project_list')

    project = result

    # Check if user is the DPM
    redirect_response = ensure_is_dpm(request, project)
    if redirect_response:
        return redirect_response

    # Check project configuration before attempting task creation
    if not project.project_incharge or not project.expected_completion_date:
        messages.warning(
            request,
            "Please complete the project configuration before creating tasks."
        )
        return redirect('projects:project_management', project_id=project_id)

    # Process form in the view
    form = ProjectTaskForm(project=project, user=request.user, data=request.POST)

    if form.is_valid():
        try:
            # Form's save method delegates to service
            form.save()
            messages.success(request, "Task created successfully")
        except ValidationError as e:
            messages.error(request, f"Error creating task: {e}")
    else:
        for field, errors in form.errors.items():
            for error in errors:
                messages.error(request, f"{field}: {error}")

    return redirect('projects:project_management', project_id=project_id)


@login_required
def update_project_configuration(request, project_id):
    """Handle project configuration updates only."""
    if request.method != 'POST':
        return redirect('projects:project_management', project_id=project_id)

    # Get project to check permissions
    success, result = ProjectService.get_project(project_id)
    if not success:
        messages.error(request, result)
        return redirect('projects:project_list')

    project = result

    # Check if user is the DPM
    redirect_response = ensure_is_dpm(request, project)
    if redirect_response:
        return redirect_response

    # Process form in the view
    form = ProjectManagementForm(request.POST, instance=project)

    if form.is_valid():
        # Pass cleaned data to service
        success, result = ProjectService.update_project_configuration(
            project_id=project_id,
            config_data=form.cleaned_data,
            dpm=request.user
        )

        if success:
            messages.success(request, "Project details updated successfully")

            # Check if delivery performance rating was updated
            if 'delivery_performance_rating' in form.changed_data:
                # Update any existing ProjectDelivery records
                from projects.models import ProjectDelivery
                ProjectDelivery.objects.filter(project=project).update(
                    delivery_performance_rating=form.cleaned_data['delivery_performance_rating']
                )

                # Recalculate metrics if there are deliveries
                deliveries = ProjectDelivery.objects.filter(project=project)
                for delivery in deliveries:
                    ProjectService.calculate_team_member_metrics(
                        delivery.project_incharge,
                        delivery.delivery_date
                    )
        else:
            messages.error(request, result)
    else:
        for field, errors in form.errors.items():
            for error in errors:
                messages.error(request, f"{field}: {error}")

    return redirect('projects:project_management', project_id=project_id)



@login_required
def task_detail(request, project_id, task_id):
    """Display task details and assignment list. Accessible by DPMs and Senior Managers."""
    # Check if user has access (DPM or Senior Manager)
    if request.user.role not in ['DPM', 'SENIOR_MANAGER']:
        messages.error(request, "Access denied. Only Project Managers and Senior Managers can view task details.")
        return redirect('home')
    
    # Get the project
    success, project_result = ProjectService.get_project(project_id)
    if not success:
        messages.error(request, project_result)
        return redirect('projects:project_list')

    project = project_result

    # Get task and assignments
    success, task_result = ProjectService.get_task_with_assignments(task_id)
    if not success:
        messages.error(request, task_result)
        return redirect('projects:project_management', project_id=project_id)

    task, assignments = task_result

    # Ensure task belongs to the correct project
    if task.project.id != project.id:
        messages.error(request, "Invalid task for this project")
        return redirect('projects:project_management', project_id=project_id)

    # Separate assignments into active and completed
    active_assignments = []
    completed_assignments = []

    for assignment in assignments:
        # Add working hours to each assignment
        assignment.working_hours = assignment.get_total_working_hours()

        if assignment.is_completed:
            completed_assignments.append(assignment)
        else:
            active_assignments.append(assignment)

    # Prepare context
    context = {
        'task': task,
        'active_assignments': active_assignments,
        'completed_assignments': completed_assignments,
        'project': project,
        'title': f'Task: {task.task_id}',
        'assignment_form': TaskAssignmentForm(),
        'update_form': TaskAssignmentUpdateForm()
    }

    return render(request, 'projects/task_detail.html', context)


@login_required
def create_task_assignment(request, project_id, task_id):
    """Handle new assignment creation."""
    if request.method != 'POST':
        return redirect('projects:task_detail', project_id=project_id, task_id=task_id)

    # Get the project and check permissions
    success, project_result = ProjectService.get_project(project_id)
    if not success:
        messages.error(request, project_result)
        return redirect('projects:project_list')

    project = project_result
    redirect_response = ensure_is_dpm(request, project)
    if redirect_response:
        return redirect_response
    
    # Additional check: ensure the user is the project's DPM
    if request.user != project.dpm:
        messages.error(request, "Access denied. Only the project's assigned DPM can create assignments.")
        return redirect('projects:task_detail', project_id=project_id, task_id=task_id)

    # Get the task to pass to the form
    success, task_result = ProjectService.get_project_task(task_id, project_id)
    if not success:
        messages.error(request, task_result)
        return redirect('projects:project_management', project_id=project_id)

    task = task_result

    # Use the form for validation and data conversion
    form = TaskAssignmentForm(data=request.POST, task=task, user=request.user)

    if form.is_valid():
        try:
            # Form's save method delegates to service
            form.save()
            messages.success(request, "Assignment created successfully")
        except ValidationError as e:
            messages.error(request, f"Error creating assignment: {e}")

    else:
        for field, errors in form.errors.items():
            for error in errors:
                messages.error(request, f"{field}: {error}")

    return redirect('projects:task_detail', project_id=project_id, task_id=task_id)


@login_required
def update_task_assignment(request, project_id, task_id, assignment_id):
    """Handle assignment updates."""
    if request.method != 'POST':
        return redirect('projects:task_detail', project_id=project_id, task_id=task_id)

    # Get the project and perform permission checks
    success, project_result = ProjectService.get_project(project_id)
    if not success:
        messages.error(request, project_result)
        return redirect('projects:project_list')

    project = project_result
    redirect_response = ensure_is_dpm(request, project)
    if redirect_response:
        return redirect_response
    
    # Additional check: ensure the user is the project's DPM
    if request.user != project.dpm:
        messages.error(request, "Access denied. Only the project's assigned DPM can update assignments.")
        return redirect('projects:task_detail', project_id=project_id, task_id=task_id)

    # First get the current assignment
    success, assignment_result = ProjectService.get_task_assignment(assignment_id)
    if not success:
        messages.error(request, assignment_result)
        return redirect('projects:task_detail', project_id=project_id, task_id=task_id)

    assignment = assignment_result

    # Add validation that assignment belongs to the correct task and project
    if assignment.task.id != task_id or assignment.task.project.id != project_id:
        messages.error(request, "Invalid assignment for this project and task")
        return redirect('projects:task_detail', project_id=project_id, task_id=task_id)

    # Use the form to validate and convert the data
    form = TaskAssignmentUpdateForm(data=request.POST, instance=assignment)

    if form.is_valid():
        # Form is valid, get the cleaned_data which has proper conversions applied
        form_data = form.cleaned_data

        # Pass the properly converted data to the service
        success, result = ProjectService.update_task_assignment(
            assignment_id=assignment_id,
            assignment_data=form_data,
            dpm=request.user
        )

        if success:
            messages.success(request, "Assignment updated successfully")
        else:
            messages.error(request, result)
    else:
        # Form validation failed
        for field, errors in form.errors.items():
            for error in errors:
                messages.error(request, f"{field}: {error}")

    return redirect('projects:task_detail', project_id=project_id, task_id=task_id)


@login_required
def dpm_task_dashboard(request):
    """
    Dashboard view for DPMs showing their projects for task management.
    This is the new Task Management tab for DPMs.
    """
    # Redirect non-DPM users to appropriate page
    if request.user.role != 'DPM':
        messages.error(request, "Access denied. This page is only for Project Managers.")
        return redirect('home')

    # Get projects for this DPM
    success, result = ProjectService.get_dpm_projects_for_task_management(request.user)

    if not success:
        messages.error(request, result)
        return redirect('home')

    projects = result

    context = {
        'projects': projects,
        'title': 'Task Management Dashboard'
    }

    return render(request, 'projects/dpm_task_dashboard.html', context)


@login_required
def team_member_dashboard(request):
    """
    Enhanced dashboard view for team members with timer functionality.
    Handles timer start/stop, manual time entry, and assignment completion.
    """
    if request.user.role != 'TEAM_MEMBER':
        messages.error(request, "This dashboard is only for team members")
        return redirect('projects:project_list')

    # Handle POST requests (timer actions)
    if request.method == 'POST':

        # Start Timer
        if 'start_timer' in request.POST:
            assignment_id = request.POST.get('assignment_id')
            success, result = ProjectService.start_timer(assignment_id, request.user)

            if success:
                messages.success(request, f"Timer started for assignment {result.assignment.assignment_id}")
            else:
                messages.error(request, result)

        # Stop Timer
        elif 'stop_timer' in request.POST:
            form = TimerStopForm(request.POST)
            if form.is_valid():
                description = form.cleaned_data['description']
                is_completed = form.cleaned_data['is_completed']

                success, result = ProjectService.stop_timer(request.user, description)

                if success:
                    messages.success(request, f"Timer stopped. Session duration: {result.get_formatted_duration()}")

                    # Mark as completed if requested
                    if is_completed:
                        success, complete_result = ProjectService.complete_assignment(
                            result.assignment.id,
                            request.user
                        )
                        if success:
                            messages.success(request, f"Assignment {complete_result.assignment_id} marked as completed!")
                        else:
                            messages.warning(request, f"Timer stopped but couldn't complete assignment: {complete_result}")
                else:
                    messages.error(request, result)

        # Add Manual Time
        elif 'add_time' in request.POST:
            assignment_id = request.POST.get('assignment_id')
            form = ManualTimeEntryForm(request.POST)

            if form.is_valid():
                success, result = ProjectService.add_manual_time(
                    assignment_id=assignment_id,
                    team_member=request.user,
                    date_worked=form.cleaned_data['date'],
                    hours=form.cleaned_data['duration_hours'],
                    minutes=form.cleaned_data['duration_minutes'],
                    description=form.cleaned_data['description'],
                    reason=form.cleaned_data['reason']
                )

                if success:
                    total_minutes = (form.cleaned_data['duration_hours'] * 60) + form.cleaned_data['duration_minutes']
                    messages.success(request, f"Added {ProjectService._format_minutes(total_minutes)} to your timesheet")

                    # Mark as completed if requested
                    if form.cleaned_data['is_completed']:
                        success, complete_result = ProjectService.complete_assignment(assignment_id, request.user)
                        if success:
                            messages.success(request, f"Assignment {complete_result.assignment_id} marked as completed!")
                        else:
                            messages.warning(request, f"Time added but couldn't complete assignment: {complete_result}")
                else:
                    messages.error(request, result)
            else:
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f"{field}: {error}")

        # Mark Completed
        elif 'mark_completed' in request.POST:
            assignment_id = request.POST.get('assignment_id')
            success, result = ProjectService.complete_assignment(assignment_id, request.user)

            if success:
                messages.success(request, f"Assignment {result.assignment_id} marked as completed!")
            else:
                messages.error(request, result)

        # Redirect to prevent form resubmission
        return redirect('projects:team_member_dashboard')

    # GET request - display dashboard
    success, result = ProjectService.get_team_member_dashboard_data(request.user)

    if not success:
        messages.error(request, result)
        return redirect('home')

    dashboard_data = result

    # Create forms for the modals
    timer_stop_form = TimerStopForm()
    time_entry_form = ManualTimeEntryForm(initial={'date': timezone.localtime(timezone.now()).date()})

    context = {
        'active_assignments': dashboard_data['active_assignments'],
        'completed_assignments': dashboard_data['completed_assignments'],
        'active_timer': dashboard_data['active_timer'],
        'elapsed_time': dashboard_data['elapsed_time'],
        'today_summary': dashboard_data['today_summary'],
        'timer_stop_form': timer_stop_form,
        'time_entry_form': time_entry_form,
        'title': 'My Tasks Dashboard',
        'now': timezone.now()  # For template comparisons
    }

    return render(request, 'projects/team_member_dashboard.html', context)


@login_required
def completed_assignments_list(request):
    """
    Display completed assignments for the current team member with optional date filtering.
    """
    if request.user.role != 'TEAM_MEMBER':
        messages.error(request, "This page is only accessible by team members")
        return redirect('projects:project_list')

    # Get date range parameters from request
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    # Parse dates if provided
    start_date = None
    end_date = None
    date_filter_active = False

    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            date_filter_active = True
        except ValueError:
            messages.warning(request, "Invalid start date format")

    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            date_filter_active = True
        except ValueError:
            messages.warning(request, "Invalid end date format")

    # Validate date range
    if start_date and end_date and start_date > end_date:
        messages.error(request, "Start date cannot be after end date")
        start_date = end_date = None
        date_filter_active = False

    # Get completed assignments with date filtering
    success, result = ProjectService.get_team_member_all_completed_assignments(
        request.user, start_date, end_date
    )

    if not success:
        messages.error(request, result)
        return redirect('projects:team_member_dashboard')

    completed_assignments = result

    # Prepare filter display text
    filter_text = "All Time"
    if date_filter_active:
        if start_date and end_date:
            filter_text = f"{start_date.strftime('%b %d, %Y')} - {end_date.strftime('%b %d, %Y')}"
        elif start_date:
            filter_text = f"From {start_date.strftime('%b %d, %Y')}"
        elif end_date:
            filter_text = f"Until {end_date.strftime('%b %d, %Y')}"

    context = {
        'completed_assignments': completed_assignments,
        'title': 'Completed Assignments',
        'total_count': len(completed_assignments),
        'start_date': start_date_str or '',
        'end_date': end_date_str or '',
        'filter_text': filter_text,
        'date_filter_active': date_filter_active
    }

    return render(request, 'projects/completed_assignments_list.html', context)


@login_required
def export_delivered_projects(request):
    """
    Export delivered projects (Final Delivery status) to CSV or XLSX format.
    Uses the same filtering logic as delivered_projects view.
    Only accessible by DPMs and Senior Managers.
    """
    # Check if user is a DPM or Senior Manager
    if request.user.role not in ['DPM', 'SENIOR_MANAGER']:
        messages.error(request, "Access denied. This page is only for Project Managers and Senior Managers.")
        return redirect('home')
    
    # Get export format from request (default to CSV)
    export_format = request.GET.get('format', 'csv').lower()
    
    # Get filter parameters from request (same as delivered_projects view)
    search_query = request.GET.get('search', '')
    status = request.GET.get('status', '')
    product = request.GET.get('product', '')
    region = request.GET.get('region', '')
    city = request.GET.get('city', '')
    dpm = request.GET.get('dpm', '')
    delivery_date_from = request.GET.get('delivery_date_from', '')
    delivery_date_to = request.GET.get('delivery_date_to', '')

    # Convert date strings to date objects if provided
    date_from = None
    date_to = None
    try:
        if delivery_date_from:
            date_from = datetime.strptime(delivery_date_from, '%Y-%m-%d').date()
    except ValueError:
        pass

    try:
        if delivery_date_to:
            date_to = datetime.strptime(delivery_date_to, '%Y-%m-%d').date()
    except ValueError:
        pass

    # Get ALL delivered projects (no pagination for export)
    success, result = ProjectService.get_project_list(
        search_query=search_query,
        status=status,
        product=product,
        region=region,
        city=city,
        dpm=dpm,
        date_from=date_from,
        date_to=date_to,
        page=1,
        items_per_page=10000,  # Large number to get all results
        project_type='delivered'
    )

    if not success:
        messages.error(request, f"Error exporting projects: {result}")
        return redirect('projects:delivered_projects')

    projects, _ = result
    projects_list = projects.object_list if hasattr(projects, 'object_list') else projects

    # Define the columns for export
    headers = [
        'HS ID', 'Opportunity ID', 'Project Name', 'Builder Name', 'City', 'Region',
        'Product', 'Product Subcategory', 'Package ID', 'Quantity', 'Purchase Date',
        'Sales Confirmation Date', 'Expected TAT', 'Account Manager', 'DPM',
        'Current Status', 'Project Incharge', 'Expected Completion Date',
        'Delivery Date', 'Delivery Performance Rating', 'Created At'
    ]

    if export_format == 'xlsx':
        return _export_delivered_to_xlsx(projects_list, headers)
    else:
        return _export_delivered_to_csv(projects_list, headers)


def _export_delivered_to_csv(projects, headers):
    """Helper function to export delivered projects to CSV format."""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="delivered_projects_{date.today().strftime("%Y%m%d")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(headers)
    
    for project in projects:
        row = [
            project.hs_id,
            project.opportunity_id,
            project.project_name,
            project.builder_name,
            project.city.name if project.city else '',
            project.city.region.name if project.city and project.city.region else '',
            project.product.name if project.product else '',
            project.product_subcategory.name if project.product_subcategory else '',
            project.package_id or '',
            project.quantity,
            project.purchase_date.strftime('%Y-%m-%d') if project.purchase_date else '',
            project.sales_confirmation_date.strftime('%Y-%m-%d') if project.sales_confirmation_date else '',
            project.expected_tat,
            project.account_manager,
            project.dpm.get_full_name() if project.dpm else '',
            project.current_status.name if project.current_status else '',
            project.project_incharge.get_full_name() if project.project_incharge else '',
            project.expected_completion_date.strftime('%Y-%m-%d') if project.expected_completion_date else '',
            project.delivery_date.strftime('%Y-%m-%d') if project.delivery_date else '',
            str(project.delivery_performance_rating) if project.delivery_performance_rating else '',
            project.created_at.strftime('%Y-%m-%d %H:%M:%S') if project.created_at else ''
        ]
        writer.writerow(row)
    
    return response


def _export_delivered_to_xlsx(projects, headers):
    """Helper function to export delivered projects to XLSX format."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Delivered Projects"
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Write data
    for row_num, project in enumerate(projects, 2):
        data = [
            project.hs_id,
            project.opportunity_id,
            project.project_name,
            project.builder_name,
            project.city.name if project.city else '',
            project.city.region.name if project.city and project.city.region else '',
            project.product.name if project.product else '',
            project.product_subcategory.name if project.product_subcategory else '',
            project.package_id or '',
            project.quantity,
            project.purchase_date.strftime('%Y-%m-%d') if project.purchase_date else '',
            project.sales_confirmation_date.strftime('%Y-%m-%d') if project.sales_confirmation_date else '',
            project.expected_tat,
            project.account_manager,
            project.dpm.get_full_name() if project.dpm else '',
            project.current_status.name if project.current_status else '',
            project.project_incharge.get_full_name() if project.project_incharge else '',
            project.expected_completion_date.strftime('%Y-%m-%d') if project.expected_completion_date else '',
            project.delivery_date.strftime('%Y-%m-%d') if project.delivery_date else '',
            str(project.delivery_performance_rating) if project.delivery_performance_rating else '',
            project.created_at.strftime('%Y-%m-%d %H:%M:%S') if project.created_at else ''
        ]
        
        for col_num, value in enumerate(data, 1):
            worksheet.cell(row=row_num, column=col_num, value=value)
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="delivered_projects_{date.today().strftime("%Y%m%d")}.xlsx"'
    
    workbook.save(response)
    return response


@login_required
def assignment_timesheet(request, assignment_id):
    """View detailed timesheet for a specific assignment."""
    # Allow team members, DPMs, and Senior Managers to access timesheets
    if request.user.role not in ['TEAM_MEMBER', 'DPM', 'SENIOR_MANAGER']:
        messages.error(request, "Access denied")
        return redirect('home')
    
    # For DPMs and Senior Managers, just verify the assignment exists (they can view any timesheet)
    if request.user.role in ['DPM', 'SENIOR_MANAGER']:
        try:
            assignment = TaskAssignment.objects.select_related('task__project').get(id=assignment_id)
        except TaskAssignment.DoesNotExist:
            messages.error(request, "Assignment not found")
            return redirect('projects:dpm_assignments_overview')

    # Handle session duration editing
    if request.method == 'POST' and 'edit_session_duration' in request.POST:
        # Get form data
        form_data = {
            'session_id': request.POST.get('session_id'),
            'duration_hours': request.POST.get('duration_hours'),
            'duration_minutes': request.POST.get('duration_minutes')
        }

        # Create and validate form
        form = EditSessionDurationForm(form_data)

        if form.is_valid():
            session_id = form.cleaned_data['session_id']
            total_minutes = form.get_total_minutes()

            # Call service method
            success, result = ProjectService.edit_session_duration(
                session_id=session_id,
                team_member=request.user,
                new_duration_minutes=total_minutes
            )

            if success:
                formatted_duration = ProjectService._format_minutes(total_minutes)
                messages.success(request, f"Session duration updated to {formatted_duration}")
            else:
                messages.error(request, f"Error updating session: {result}")
        else:
            # Handle form validation errors
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")

        return redirect('projects:assignment_timesheet', assignment_id=assignment_id)

    # Get timesheet data using service layer (no date filtering)
    # For DPMs and Senior Managers, pass the actual team member instead of the current user
    if request.user.role in ['DPM', 'SENIOR_MANAGER']:
        # assignment is already fetched above for DPM/Senior Manager access check
        team_member = assignment.assigned_to
    else:
        team_member = request.user
        
    success, result = ProjectService.get_assignment_timesheet_data(
        assignment_id, team_member
    )

    if not success:
        messages.error(request, result)
        if request.user.role in ['DPM', 'SENIOR_MANAGER']:
            return redirect('projects:dpm_assignments_overview')
        else:
            return redirect('projects:team_member_dashboard')

    timesheet_data = result

    # Determine the correct back URL based on user role
    from urllib.parse import unquote
    original_referer = unquote(request.GET.get('original_referer', ''))
    
    # DEBUG: Add temporary debugging
    print(f"DEBUG assignment_timesheet: original_referer from URL = '{original_referer}'")
    print(f"DEBUG assignment_timesheet: HTTP_REFERER = '{request.META.get('HTTP_REFERER', '')}'")
    
    if request.user.role in ['DPM', 'SENIOR_MANAGER']:
        # First check if coming from team roster or task detail
        from_param = request.GET.get('from', '')
        if from_param == 'team_roster':
            # Get the assigned team member for the back URL
            assigned_team_member = timesheet_data['assignment'].assigned_to
            back_url = f"/projects/team-roster/daily/?team_member={assigned_team_member.id}"
            back_text = 'Back to Roster'
            back_is_full_url = True
        elif from_param == 'task_detail':
            # Get project and task IDs for the back URL
            project_id = timesheet_data['assignment'].task.project.id
            task_id = timesheet_data['assignment'].task.id
            back_url = f"/projects/{project_id}/tasks/{task_id}/"
            back_text = 'Back to Task'
            back_is_full_url = True
        else:
            # Check if we have a preserved original_referer from quality rating flow
            if original_referer:
                referer = original_referer
            else:
                # Extract filter parameters from HTTP_REFERER if coming from assignments overview
                referer = request.META.get('HTTP_REFERER', '')
                
            if '/projects/tasks/assignments/' in referer and '?' in referer:
                # Extract query parameters from referer URL
                from urllib.parse import urlparse, parse_qs
                parsed_url = urlparse(referer)
                query_params = parse_qs(parsed_url.query)
                
                # Build filter parameters
                filter_params = []
                for param in ['assignment_status', 'team_member', 'dpm', 'project', 'start_date', 'end_date']:
                    if param in query_params and query_params[param][0]:
                        filter_params.append(f"{param}={query_params[param][0]}")
                
                if filter_params:
                    # Reconstruct the assignments overview URL with filters
                    back_url = f"/projects/tasks/assignments/?{'&'.join(filter_params)}"
                    back_is_full_url = True
                else:
                    # Default assignments overview
                    back_url = 'projects:dpm_assignments_overview'
                    back_is_full_url = False
            else:
                # Default assignments overview
                back_url = 'projects:dpm_assignments_overview'
                back_is_full_url = False
                
            back_text = 'Back to Assignments'
    else:
        # Check if coming from daily roster, team roster, or task detail
        from_param = request.GET.get('from', '')
        if from_param == 'daily_roster':
            back_url = 'projects:daily_roster'
            back_text = 'Back to Roster'
            back_is_full_url = False
        elif from_param == 'team_roster':
            back_url = 'projects:team_member_daily_roster'
            back_text = 'Back to Roster'
            back_is_full_url = False
        elif from_param == 'task_detail':
            # Get project and task IDs for the back URL
            project_id = timesheet_data['assignment'].task.project.id
            task_id = timesheet_data['assignment'].task.id
            back_url = f"/projects/{project_id}/tasks/{task_id}/"
            back_text = 'Back to Task'
            back_is_full_url = True
        else:
            back_url = 'projects:team_member_dashboard'
            back_text = 'Back to Dashboard'
            back_is_full_url = False

    # Prepare quality rating options for DPMs (only for the project's DPM)
    quality_rating_options = []
    if (request.user.role == 'DPM' and assignment.is_completed and 
        request.user == timesheet_data['assignment'].task.project.dpm):
        quality_rating_options = [
            {
                'value': 1.0,
                'stars': [1],
                'description': 'Poor'
            },
            {
                'value': 2.0,
                'stars': [1, 2],
                'description': 'Fair'
            },
            {
                'value': 3.0,
                'stars': [1, 2, 3],
                'description': 'Good'
            },
            {
                'value': 4.0,
                'stars': [1, 2, 3, 4],
                'description': 'Very Good'
            },
            {
                'value': 5.0,
                'stars': [1, 2, 3, 4, 5],
                'description': 'Excellent'
            }
        ]

    # Determine the original referer to pass to the template
    # Priority: URL parameter > current HTTP_REFERER (only if not from quality rating)
    template_original_referer = original_referer
    if not template_original_referer:
        current_referer = request.META.get('HTTP_REFERER', '')
        # Only use current referer if it's from assignments overview, not from quality rating
        if '/projects/tasks/assignments/' in current_referer:
            template_original_referer = current_referer

    context = {
        'assignment': timesheet_data['assignment'],
        'daily_totals': timesheet_data['daily_totals'],
        'sessions': timesheet_data['sessions'],
        'assignment_summary': timesheet_data['assignment_summary'],
        'back_url': back_url,
        'back_text': back_text,
        'back_is_full_url': back_is_full_url,
        'quality_rating_options': quality_rating_options,
        'original_referer': template_original_referer,
        'title': f'Timesheet: {timesheet_data["assignment"].assignment_id}'
    }

    return render(request, 'projects/assignment_timesheet.html', context)

@login_required
def daily_roster(request):
    """View team member's daily time breakdown."""
    if request.user.role != 'TEAM_MEMBER':
        messages.error(request, "Access denied")
        return redirect('home')

    # Get filter parameters directly from request GET, with defaults
    try:
        default_date = timezone.localtime(timezone.now()).date()
        date_str = request.GET.get('date')
        selected_date = datetime.strptime(date_str, '%Y-%m-%d').date() if date_str else default_date
    except (ValueError, TypeError):
        selected_date = default_date

    show_week = request.GET.get('week_view') == 'on'

    # Create a filter form instance, initialized with the current values for rendering
    filter_form = DailyRosterFilterForm(initial={
        'date': selected_date,
        'week_view': show_week,
    })

    # Use the determined values to fetch data from the service layer
    success, roster_data = ProjectService.get_daily_roster_data(
        request.user, selected_date, show_week
    )

    if not success:
        messages.error(request, roster_data)
        roster_data = {
            'daily_totals': [],
            'daily_rosters': {},
            'misc_hours_entries': [],
            'date_range': 'Error loading data',
            'total_formatted': '00:00',
            'assignment_minutes': 0,
            'misc_minutes': 0,
        }
    
    # Calculate daily summaries if in week view
    daily_summaries = {}
    if show_week and success:
        # Group assignment and misc totals by day
        day_totals_map = {}
        for dt in roster_data['daily_totals']:
            day_key = dt.date_worked
            day_totals_map.setdefault(day_key, 0)
            day_totals_map[day_key] += dt.total_minutes

        for me in roster_data['misc_hours_entries']:
            day_key = me.date
            day_totals_map.setdefault(day_key, 0)
            day_totals_map[day_key] += me.duration_minutes

        # Format totals into HH:MM strings
        for day, total_minutes in day_totals_map.items():
            daily_summaries[day] = ProjectService._format_minutes(total_minutes)

    context = {
        'daily_totals': roster_data['daily_totals'],
        'daily_rosters': roster_data['daily_rosters'],
        'misc_hours_entries': roster_data['misc_hours_entries'],
        'filter_form': filter_form,
        'selected_date': selected_date,
        'show_week': show_week,
        'date_range': roster_data['date_range'],
        'total_formatted': roster_data['total_formatted'],
        'assignment_minutes': roster_data['assignment_minutes'],
        'misc_minutes': roster_data['misc_minutes'],
        'daily_summaries': daily_summaries,
        'misc_activity_type_choices': MiscHours.ACTIVITY_TYPE_CHOICES,
        'title': f'Daily Roster - {roster_data["date_range"]}'
    }

    return render(request, 'projects/daily_roster.html', context)

@login_required
def monthly_roster(request, year=None, month=None):
    """
    Display monthly roster for team member.
    Shows calendar view with daily status and hours.
    Now includes misc hours functionality.
    """
    if request.user.role != 'TEAM_MEMBER':
        messages.error(request, "This page is only for team members")
        return redirect('home')

    # Handle misc hours addition
    if request.method == 'POST' and 'add_misc_hours' in request.POST:
        form = AddMiscHoursForm(request.POST)

        if form.is_valid():
            success, result = ProjectService.add_misc_hours(
                team_member=request.user,
                work_date=form.cleaned_data['date'],
                activity=form.cleaned_data['activity'],
                duration_hours=form.cleaned_data['duration_hours'],
                duration_minutes=form.cleaned_data['duration_minutes'],
                activity_type=form.cleaned_data['activity_type']
            )

            if success:
                total_minutes = form.get_total_minutes()
                formatted_duration = ProjectService._format_minutes(total_minutes)
                messages.success(request, f"Added {formatted_duration} of misc work: {form.cleaned_data['activity']}")
            else:
                messages.error(request, f"Error adding misc hours: {result}")
        else:
            # Handle form validation errors
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")

        # Redirect to the same month after processing
        if year and month:
            return redirect('projects:roster_date', year=year, month=month)  # FIXED: Changed from 'monthly_roster_date'
        else:
            return redirect('projects:roster')  # FIXED: Changed from 'monthly_roster'

    # Default to current month if not specified
    if not year or not month:
        today = date.today()
        year, month = today.year, today.month

    # Get monthly roster data using optimized method
    success, result = ProjectService.get_monthly_roster_optimized(request.user, year, month)

    if not success:
        messages.error(request, result)
        return redirect('projects:team_member_dashboard')

    monthly_data = result

    # Calculate navigation dates
    current_date = date(year, month, 1)
    prev_month = current_date - timedelta(days=1)
    next_month_day = current_date.replace(day=28) + timedelta(days=4)
    next_month = next_month_day - timedelta(days=next_month_day.day-1)

    # Create misc hours form for the modal
    misc_hours_form = AddMiscHoursForm()

    context = {
        'monthly_data': monthly_data,
        'current_date': current_date,
        'prev_month': prev_month,
        'next_month': next_month,
        'misc_hours_form': misc_hours_form,
        'title': f'Roster - {monthly_data["month_name"]} {year}'
    }

    return render(request, 'projects/monthly_roster.html', context)

@login_required
def update_roster_day(request):
    """
    Update roster status for a specific day.
    """
    if request.user.role != 'TEAM_MEMBER':
        messages.error(request, "Access denied")
        return redirect('home')

    if request.method != 'POST':
        return redirect('projects:roster')  # FIXED: Changed from 'monthly_roster'

    try:
        # Get form data
        date_str = request.POST.get('date')
        new_status = request.POST.get('status')
        notes = request.POST.get('notes', '')

        # Parse date
        from datetime import datetime
        work_date = datetime.strptime(date_str, '%Y-%m-%d').date()

        # Get existing roster to preserve misc hours and description
        roster, _ = ProjectService.get_or_create_daily_roster(request.user, work_date)

        # Update roster using service
        success, result = ProjectService.update_roster_status(
            team_member=request.user,
            date=work_date,
            new_status=new_status,
            misc_hours=roster.misc_hours,
            misc_description=roster.misc_description,
            notes=notes
        )

        if success:
            messages.success(request, f"Status updated for {work_date.strftime('%B %d, %Y')}")
        else:
            messages.error(request, f"Error updating status: {result}")

    except Exception as e:
        messages.error(request, f"Invalid date format: {str(e)}")

    # Redirect back to the same month
    try:
        work_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        return redirect('projects:roster_date', year=work_date.year, month=work_date.month)
    except:
        return redirect('projects:roster')


@login_required
def update_quality_rating(request, project_id, task_id, assignment_id):
    """
    Dedicated view for updating only the quality rating of completed assignments.
    Bypasses the complex TaskAssignmentUpdateForm validation.
    """
    if request.method != 'POST':
        return redirect('projects:task_detail', project_id=project_id, task_id=task_id)

    # Get the project and check permissions
    success, project_result = ProjectService.get_project(project_id)
    if not success:
        messages.error(request, project_result)
        return redirect('projects:project_list')

    project = project_result
    redirect_response = ensure_is_dpm(request, project)
    if redirect_response:
        return redirect_response
    
    # Additional check: ensure the user is the project's DPM
    if request.user != project.dpm:
        messages.error(request, "Access denied. Only the project's assigned DPM can rate assignments.")
        return redirect('projects:task_detail', project_id=project_id, task_id=task_id)

    # Get the assignment
    success, assignment_result = ProjectService.get_task_assignment(assignment_id)
    if not success:
        messages.error(request, assignment_result)
        return redirect('projects:task_detail', project_id=project_id, task_id=task_id)

    assignment = assignment_result

    # Validate assignment belongs to correct task and project
    if assignment.task.id != task_id or assignment.task.project.id != project_id:
        messages.error(request, "Invalid assignment for this project and task")
        return redirect('projects:task_detail', project_id=project_id, task_id=task_id)

    # Validate assignment is completed
    if not assignment.is_completed:
        messages.error(request, "Can only rate quality of completed assignments")
        return redirect('projects:task_detail', project_id=project_id, task_id=task_id)

    # Get and validate quality rating
    quality_rating = request.POST.get('quality_rating')
    if not quality_rating:
        messages.error(request, "Please select a quality rating")
        return redirect('projects:task_detail', project_id=project_id, task_id=task_id)

    # Get optional comments
    quality_rating_comments = request.POST.get('quality_rating_comments', '').strip()

    try:
        # Convert to decimal and validate range
        rating_value = float(quality_rating)
        if rating_value < 1.0 or rating_value > 5.0:
            messages.error(request, "Quality rating must be between 1.0 and 5.0")
            return redirect('projects:task_detail', project_id=project_id, task_id=task_id)

        # Update quality rating and comments
        assignment.quality_rating = rating_value
        assignment.quality_rating_comments = quality_rating_comments
        assignment.save()

        messages.success(request, f"Quality rating updated to {rating_value}/5 for assignment {assignment.assignment_id}")

    except (ValueError, TypeError):
        messages.error(request, "Invalid quality rating value")

    return redirect('projects:task_detail', project_id=project_id, task_id=task_id)


@login_required
def update_quality_rating_timesheet(request, assignment_id):
    """
    Update quality rating for a completed assignment directly from the timesheet.
    Only accessible by DPMs and Senior Managers.
    """
    if request.user.role != 'DPM':
        messages.error(request, "Access denied. Only Project Managers can rate assignments.")
        return redirect('home')

    if request.method != 'POST':
        messages.error(request, "Invalid request method.")
        return redirect('projects:assignment_timesheet', assignment_id=assignment_id)

    try:
        assignment = TaskAssignment.objects.select_related(
            'task__project', 'assigned_to'
        ).get(id=assignment_id)
    except TaskAssignment.DoesNotExist:
        messages.error(request, "Assignment not found.")
        return redirect('projects:dpm_assignments_overview')

    # Validate assignment is completed
    if not assignment.is_completed:
        messages.error(request, "Can only rate quality of completed assignments")
        return redirect('projects:assignment_timesheet', assignment_id=assignment_id)

    # Preserve the original referer for filter preservation
    original_referer = request.POST.get('original_referer', '')
    
    # DEBUG: Add temporary debugging
    print(f"DEBUG quality_rating: original_referer from POST = '{original_referer}'")
    print(f"DEBUG quality_rating: HTTP_REFERER = '{request.META.get('HTTP_REFERER', '')}'")

    # Handle clear rating request
    if request.POST.get('clear_rating'):
        assignment.quality_rating = None
        assignment.save()
        messages.success(request, f"Quality rating cleared for assignment {assignment.assignment_id}")
        
        # Redirect back to timesheet with preserved referer
        if original_referer:
            encoded_referer = quote(original_referer, safe='')
            redirect_url = f"{reverse('projects:assignment_timesheet', args=[assignment_id])}?original_referer={encoded_referer}"
            print(f"DEBUG quality_rating: redirecting to = '{redirect_url}'")
            return redirect(redirect_url)
        return redirect('projects:assignment_timesheet', assignment_id=assignment_id)

    # Get and validate quality rating
    quality_rating = request.POST.get('quality_rating')
    if not quality_rating:
        messages.error(request, "Please select a quality rating")
        
        # Redirect back to timesheet with preserved referer
        if original_referer:
            encoded_referer = quote(original_referer, safe='')
            redirect_url = f"{reverse('projects:assignment_timesheet', args=[assignment_id])}?original_referer={encoded_referer}"
            print(f"DEBUG quality_rating: redirecting to = '{redirect_url}'")
            return redirect(redirect_url)
        return redirect('projects:assignment_timesheet', assignment_id=assignment_id)

    # Get optional comments
    quality_rating_comments = request.POST.get('quality_rating_comments', '').strip()

    try:
        # Convert to decimal and validate range
        rating_value = float(quality_rating)
        if rating_value < 1.0 or rating_value > 5.0:
            messages.error(request, "Quality rating must be between 1.0 and 5.0")
            
            # Redirect back to timesheet with preserved referer
            if original_referer:
                encoded_referer = quote(original_referer, safe='')
                redirect_url = f"{reverse('projects:assignment_timesheet', args=[assignment_id])}?original_referer={encoded_referer}"
                print(f"DEBUG quality_rating: redirecting to = '{redirect_url}'")
                return redirect(redirect_url)
            return redirect('projects:assignment_timesheet', assignment_id=assignment_id)

        # Update the quality rating and comments
        assignment.quality_rating = rating_value
        assignment.quality_rating_comments = quality_rating_comments
        assignment.save()

        # Create success message with rating description
        rating_descriptions = {
            1.0: 'Poor', 2.0: 'Fair', 3.0: 'Good', 4.0: 'Very Good', 5.0: 'Excellent'
        }
        description = rating_descriptions.get(rating_value, 'Unknown')
        
        messages.success(request, 
            f"Quality rating set to {rating_value}/5 ({description}) for assignment {assignment.assignment_id}")

    except (ValueError, TypeError):
        messages.error(request, "Invalid quality rating value")

    # Redirect back to timesheet with preserved referer
    if original_referer:
        encoded_referer = quote(original_referer, safe='')
        redirect_url = f"{reverse('projects:assignment_timesheet', args=[assignment_id])}?original_referer={encoded_referer}"
        print(f"DEBUG quality_rating: redirecting to = '{redirect_url}'")
        return redirect(redirect_url)
    return redirect('projects:assignment_timesheet', assignment_id=assignment_id)


@login_required
def my_projects(request):
    """
    Display all projects where the current user is the project in-charge.
    This is a read-only view for team members.
    """
    # Check if user is a team member
    if request.user.role != 'TEAM_MEMBER':
        messages.warning(request, "This view is only available for team members.")
        return redirect('home')

    # Get projects where user is project in-charge
    success, result = ProjectService.get_team_member_projects(request.user)

    if not success:
        messages.error(request, result)
        return redirect('projects:team_member_dashboard')

    projects_data = result

    # Separate pipeline and delivered projects
    pipeline_projects = [p for p in projects_data if p['stats']['is_pipeline']]
    delivered_projects = [p for p in projects_data if p['stats']['is_delivered']]

    context = {
        'projects_data': projects_data,
        'pipeline_projects': pipeline_projects,
        'delivered_projects': delivered_projects,
        'title': 'My Projects',
        'total_projects': len(projects_data),
        'is_read_only': True  # Flag to ensure read-only display
    }

    return render(request, 'projects/my_projects.html', context)


@login_required
def dpm_assignments_overview(request):
    """
    Display all task assignments for management roles with filtering capabilities.
    Shows both active and completed assignments with timesheet links.
    """
    # Check if user has management access (DPM, VIDEO_PM, or SENIOR_MANAGER)
    redirect_response = ensure_has_management_access(request)
    if redirect_response:
        return redirect_response

    # Handle AJAX request for dynamic filtering
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        if request.POST.get('get_filter_options'):
            # Get filter parameters from POST
            assignment_status = request.POST.get('assignment_status', 'all')
            team_member_id = request.POST.get('team_member', '')
            dpm_id = request.POST.get('dpm', '')
            project_id = request.POST.get('project', '')
            
            # Convert to objects if provided
            team_member = None
            dpm = None
            project = None
            try:
                if team_member_id:
                    team_member = User.objects.get(id=team_member_id)
                if dpm_id:
                    dpm = User.objects.get(id=dpm_id)
                if project_id:
                    project = Project.objects.get(id=project_id)
            except (User.DoesNotExist, Project.DoesNotExist):
                pass
            
            # Get assignments for projects (excluding project filter)
            success_projects, result_projects = ProjectService.get_dpm_all_task_assignments(
                assignment_status=assignment_status,
                team_member=team_member,
                project=None,  # Don't filter by project for project choices
                dpm=dpm,
                start_date=None,
                end_date=None
            )
            
            # Get assignments for team members (excluding team member filter)
            success_members, result_members = ProjectService.get_dpm_all_task_assignments(
                assignment_status=assignment_status,
                team_member=None,  # Don't filter by team member for team member choices
                project=project,
                dpm=dpm,
                start_date=None,
                end_date=None
            )
            
            response_data = {}
            
            # Get unique projects
            if success_projects:
                project_ids = set()
                for assignment in result_projects:
                    project_ids.add(assignment.task.project.id)
                
                projects = Project.objects.filter(id__in=project_ids).order_by('project_name')
                response_data['projects'] = [
                    {'id': p.id, 'name': p.project_name}
                    for p in projects
                ]
            else:
                response_data['projects'] = []
            
            # Get unique team members
            if success_members:
                member_ids = set()
                for assignment in result_members:
                    member_ids.add(assignment.assigned_to.id)
                
                members = User.objects.filter(id__in=member_ids, role='TEAM_MEMBER').order_by('first_name', 'last_name', 'username')
                response_data['team_members'] = [
                    {
                        'id': m.id, 
                        'name': m.get_full_name() if m.get_full_name().strip() else m.username
                    }
                    for m in members
                ]
            else:
                response_data['team_members'] = []
                
            return JsonResponse(response_data)

    # Regular GET request handling
    # Check if this is a fresh load (no GET parameters) and redirect to defaults
    if not request.GET:
        from datetime import date, timedelta
        today = date.today()
        start_default = today - timedelta(days=15)
        
        # Redirect with default parameters
        return redirect(f"{request.path}?assignment_status=active&start_date={start_default}&end_date={today}")
    
    # Create filter form with defaults enabled
    filter_form = TaskAssignmentFilterForm(data=request.GET, use_defaults=True)
    
    # Default parameters
    assignment_status = 'active'  # Changed default from 'all' to 'active'
    team_member = None
    project = None
    dpm = None
    start_date = None
    end_date = None
    
    # Apply filters if form is valid
    if filter_form.is_valid():
        assignment_status = filter_form.cleaned_data.get('assignment_status', 'active')
        team_member = filter_form.cleaned_data.get('team_member')
        project = filter_form.cleaned_data.get('project')
        dpm = filter_form.cleaned_data.get('dpm')
        start_date = filter_form.cleaned_data.get('start_date')
        end_date = filter_form.cleaned_data.get('end_date')
    
    # Get assignments using service layer
    success, result = ProjectService.get_dpm_all_task_assignments(
        assignment_status=assignment_status,
        team_member=team_member,
        project=project,
        dpm=dpm,
        start_date=start_date,
        end_date=end_date
    )
    
    if not success:
        messages.error(request, result)
        return redirect('projects:dpm_task_dashboard')
    
    assignments = result
    
    # Create summary statistics
    total_assignments = len(assignments)
    active_count = sum(1 for a in assignments if not a.is_completed)
    completed_count = sum(1 for a in assignments if a.is_completed)
    
    # Create date range display text for UI
    date_range_text = ""
    if start_date or end_date:
        if assignment_status == 'completed':
            date_field_name = "completion date"
        else:
            date_field_name = "assigned date"
            
        if start_date and end_date:
            date_range_text = f"From {start_date.strftime('%b %d, %Y')} to {end_date.strftime('%b %d, %Y')} ({date_field_name})"
        elif start_date:
            date_range_text = f"From {start_date.strftime('%b %d, %Y')} onwards ({date_field_name})"
        elif end_date:
            date_range_text = f"Up to {end_date.strftime('%b %d, %Y')} ({date_field_name})"
    
    # Calculate default dates for template use
    from datetime import date, timedelta
    today = date.today()
    default_start_date = today - timedelta(days=15)
    
    context = {
        'assignments': assignments,
        'filter_form': filter_form,
        'total_assignments': total_assignments,
        'active_count': active_count,
        'completed_count': completed_count,
        'assignment_status': assignment_status,
        'date_range_text': date_range_text,
        'title': 'Task Assignments Overview',
        'default_start_date': default_start_date.strftime('%Y-%m-%d'),
        'default_end_date': today.strftime('%Y-%m-%d')
    }
    
    return render(request, 'projects/dpm_assignments_overview.html', context)

@login_required
def assignment_graph_view(request):
    """
    Display assignment workload graph showing projected hours allocated to each team member.
    Shows graphical representation of active assignment workload distribution.
    """
    # Check if user has management access (DPM, VIDEO_PM, or SENIOR_MANAGER)
    redirect_response = ensure_has_management_access(request)
    if redirect_response:
        return redirect_response

    # Handle AJAX request for dynamic filtering (same as overview)
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        if request.POST.get('get_filter_options'):
            # Get filter parameters from POST
            assignment_status = request.POST.get('assignment_status', 'active')  # Default to active for graph
            team_member_id = request.POST.get('team_member', '')
            dpm_id = request.POST.get('dpm', '')
            project_id = request.POST.get('project', '')
            
            # Convert to objects if provided
            team_member = None
            dpm = None
            project = None
            try:
                if team_member_id:
                    team_member = User.objects.get(id=team_member_id)
                if dpm_id:
                    dpm = User.objects.get(id=dpm_id)
                if project_id:
                    project = Project.objects.get(id=project_id)
            except (User.DoesNotExist, Project.DoesNotExist):
                pass
            
            # Get assignments for projects (excluding project filter)
            success_projects, result_projects = ProjectService.get_dpm_all_task_assignments(
                assignment_status=assignment_status,
                team_member=team_member,
                project=None,  # Don't filter by project for project choices
                dpm=dpm,
                start_date=None,
                end_date=None
            )
            
            # Get assignments for team members (excluding team member filter)
            success_members, result_members = ProjectService.get_dpm_all_task_assignments(
                assignment_status=assignment_status,
                team_member=None,  # Don't filter by team member for team member choices
                project=project,
                dpm=dpm,
                start_date=None,
                end_date=None
            )
            
            response_data = {}
            
            # Get unique projects
            if success_projects:
                project_ids = set()
                for assignment in result_projects:
                    project_ids.add(assignment.task.project.id)
                
                projects = Project.objects.filter(id__in=project_ids).order_by('project_name')
                response_data['projects'] = [
                    {'id': p.id, 'name': p.project_name}
                    for p in projects
                ]
            else:
                response_data['projects'] = []
            
            # Get unique team members
            if success_members:
                member_ids = set()
                for assignment in result_members:
                    member_ids.add(assignment.assigned_to.id)
                
                members = User.objects.filter(id__in=member_ids, role='TEAM_MEMBER').order_by('first_name', 'last_name', 'username')
                response_data['team_members'] = [
                    {
                        'id': m.id, 
                        'name': m.get_full_name() if m.get_full_name().strip() else m.username
                    }
                    for m in members
                ]
            else:
                response_data['team_members'] = []
                
            return JsonResponse(response_data)

    # Check if it's an AJAX request for chart data
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        if request.POST.get('get_chart_data'):
            # Get filter parameters
            assignment_status = request.POST.get('assignment_status', 'active')
            team_member_id = request.POST.get('team_member', '')
            dpm_id = request.POST.get('dpm', '')
            project_id = request.POST.get('project', '')
            start_date_str = request.POST.get('start_date', '')
            end_date_str = request.POST.get('end_date', '')
            
            # Convert to objects if provided
            team_member = None
            dpm = None
            project = None
            start_date = None
            end_date = None
            
            try:
                if team_member_id:
                    team_member = User.objects.get(id=team_member_id)
                if dpm_id:
                    dpm = User.objects.get(id=dpm_id)
                if project_id:
                    project = Project.objects.get(id=project_id)
                if start_date_str:
                    start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                if end_date_str:
                    end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            except (User.DoesNotExist, Project.DoesNotExist, ValueError):
                pass
            
            # Get assignments using service layer
            success, assignments = ProjectService.get_dpm_all_task_assignments(
                assignment_status=assignment_status,
                team_member=team_member,
                project=project,
                dpm=dpm,
                start_date=start_date,
                end_date=end_date
            )
            
            if not success:
                return JsonResponse({'error': assignments}, status=400)
            
            # Aggregate data by team member
            member_workload = {}
            total_assignments = 0
            
            for assignment in assignments:
                # Only include assignments for current team members
                if assignment.assigned_to.role != 'TEAM_MEMBER':
                    continue
                    
                member_name = assignment.assigned_to.get_full_name() or assignment.assigned_to.username
                projected_hours = assignment.projected_hours or 0
                projected_hours_decimal = projected_hours / 60.0  # Convert minutes to hours
                
                # Calculate worked hours for this assignment
                worked_minutes = assignment.daily_totals.aggregate(
                    total=Sum('total_minutes')
                )['total'] or 0
                worked_hours_decimal = worked_minutes / 60.0
                
                if member_name not in member_workload:
                    member_workload[member_name] = {
                        'hours': 0,
                        'worked_hours': 0,
                        'assignments': 0,
                        'username': assignment.assigned_to.username
                    }
                
                member_workload[member_name]['hours'] += projected_hours_decimal
                member_workload[member_name]['worked_hours'] += worked_hours_decimal
                member_workload[member_name]['assignments'] += 1
                total_assignments += 1
            
            # Sort by hours (descending)
            sorted_workload = sorted(member_workload.items(), key=lambda x: x[1]['hours'], reverse=True)
            
            # Prepare chart data
            labels = [item[0] for item in sorted_workload]
            hours_data = [round(item[1]['hours'], 1) for item in sorted_workload]
            worked_hours_data = [round(item[1]['worked_hours'], 1) for item in sorted_workload]
            remaining_hours_data = [round(max(0, item[1]['hours'] - item[1]['worked_hours']), 1) for item in sorted_workload]
            assignments_data = [item[1]['assignments'] for item in sorted_workload]
            
            chart_data = {
                'labels': labels,
                'hours': hours_data,
                'worked_hours': worked_hours_data,
                'remaining_hours': remaining_hours_data,
                'assignments': assignments_data,
                'total_members': len(member_workload),
                'total_assignments': total_assignments,
                'total_hours': round(sum(hours_data), 1),
                'total_worked_hours': round(sum(worked_hours_data), 1),
                'avg_hours_per_member': round(sum(hours_data) / len(member_workload), 1) if member_workload else 0
            }
            
            return JsonResponse(chart_data)

    # Regular GET request handling
    # Create filter form, excluding previous filter parameters to ensure graph view uses defaults
    graph_filters = request.GET.copy()
    
    # Remove previous filter parameters so graph view uses its own defaults
    prev_params = ['prev_assignment_status', 'prev_team_member', 'prev_dpm', 'prev_project', 'prev_start_date', 'prev_end_date']
    for param in prev_params:
        graph_filters.pop(param, None)
    
    filter_form = TaskAssignmentFilterForm(data=graph_filters)
    
    # Default to active assignments for graph view
    assignment_status = 'active'
    team_member = None
    project = None
    dpm = None
    start_date = None
    end_date = None
    
    # Apply filters if form is valid
    if filter_form.is_valid():
        assignment_status = filter_form.cleaned_data.get('assignment_status', 'active')
        team_member = filter_form.cleaned_data.get('team_member')
        project = filter_form.cleaned_data.get('project')
        dpm = filter_form.cleaned_data.get('dpm')
        start_date = filter_form.cleaned_data.get('start_date')
        end_date = filter_form.cleaned_data.get('end_date')
    
    # Get assignments using service layer
    success, result = ProjectService.get_dpm_all_task_assignments(
        assignment_status=assignment_status,
        team_member=team_member,
        project=project,
        dpm=dpm,
        start_date=start_date,
        end_date=end_date
    )
    
    if not success:
        messages.error(request, result)
        return redirect('projects:dpm_task_dashboard')
    
    assignments = result
    
    # Aggregate data by team member for initial load
    member_workload = {}
    total_assignments = len(assignments)
    
    for assignment in assignments:
        # Only include assignments for current team members
        if assignment.assigned_to.role != 'TEAM_MEMBER':
            continue
            
        member_name = assignment.assigned_to.get_full_name() or assignment.assigned_to.username
        projected_hours = assignment.projected_hours or 0
        projected_hours_decimal = projected_hours / 60.0  # Convert minutes to hours
        
        # Calculate worked hours for this assignment
        worked_minutes = assignment.daily_totals.aggregate(
            total=Sum('total_minutes')
        )['total'] or 0
        worked_hours_decimal = worked_minutes / 60.0
        
        if member_name not in member_workload:
            member_workload[member_name] = {
                'hours': 0,
                'worked_hours': 0,
                'assignments': 0,
                'username': assignment.assigned_to.username
            }
        
        member_workload[member_name]['hours'] += projected_hours_decimal
        member_workload[member_name]['worked_hours'] += worked_hours_decimal
        member_workload[member_name]['assignments'] += 1
    
    # Sort by hours (descending)
    sorted_workload = sorted(member_workload.items(), key=lambda x: x[1]['hours'], reverse=True)
    
    # Prepare initial chart data
    worked_hours_list = [round(item[1]['worked_hours'], 1) for item in sorted_workload]
    hours_list = [round(item[1]['hours'], 1) for item in sorted_workload]
    
    initial_chart_data = {
        'labels': [item[0] for item in sorted_workload],
        'hours': hours_list,
        'worked_hours': worked_hours_list,
        'remaining_hours': [round(max(0, hours_list[i] - worked_hours_list[i]), 1) for i in range(len(hours_list))],
        'assignments': [item[1]['assignments'] for item in sorted_workload],
        'total_members': len(member_workload),
        'total_assignments': total_assignments,
        'total_hours': round(sum(hours_list), 1),
        'total_worked_hours': round(sum(worked_hours_list), 1),
        'avg_hours_per_member': round(sum(hours_list) / len(member_workload), 1) if member_workload else 0
    }
    
    # Create date range display text for UI
    date_range_text = ""
    if start_date or end_date:
        if assignment_status == 'completed':
            date_field_name = "completion date"
        else:
            date_field_name = "assigned date"
            
        if start_date and end_date:
            date_range_text = f"From {start_date.strftime('%b %d, %Y')} to {end_date.strftime('%b %d, %Y')} ({date_field_name})"
        elif start_date:
            date_range_text = f"From {start_date.strftime('%b %d, %Y')} onwards ({date_field_name})"
        elif end_date:
            date_range_text = f"Up to {end_date.strftime('%b %d, %Y')} ({date_field_name})"
    
    context = {
        'filter_form': filter_form,
        'assignment_status': assignment_status,
        'date_range_text': date_range_text,
        'initial_chart_data': json.dumps(initial_chart_data),
        'workload_data': sorted_workload,
        'title': 'Assignment Workload Graph'
    }
    
    return render(request, 'projects/assignment_graph_view.html', context)

@login_required
def edit_misc_hours(request, misc_hours_id):
    """
    Handle editing of miscellaneous hours entries via AJAX.
    """
    misc_entry = get_object_or_404(MiscHours, id=misc_hours_id, team_member=request.user)
    
    if request.method == 'POST':
        form = EditMiscHoursForm(request.POST, instance=misc_entry)
        if form.is_valid():
            form.save()
            messages.success(request, "Miscellaneous hours entry updated successfully.")
            return JsonResponse({'success': True})
        else:
            return JsonResponse({'success': False, 'errors': form.errors.get_json_data()})
    
    return JsonResponse({'success': False, 'message': 'Invalid request method.'}, status=400)


# Team Roster Views (for DPMs)
@login_required
def team_roster_list(request):
    """
    Display list of all team members with enhanced cards showing previous day snapshots.
    """
    # Check if user has management access (DPM, VIDEO_PM, or SENIOR_MANAGER)
    redirect_response = ensure_has_management_access(request)
    if redirect_response:
        return redirect_response
    
    # Get all team members (excluding DPMs)
    team_members = User.objects.filter(role='TEAM_MEMBER').order_by('first_name', 'last_name')
    
    # Get current month and date info
    today = date.today()
    
    # Get previous day snapshots for all team members (optimized bulk query)
    previous_day_snapshots = ProjectService.get_bulk_previous_day_snapshots(team_members, today)
    
    # Build team member data with both monthly summary and previous day snapshot
    team_member_data = []
    
    for member in team_members:
        # Get monthly summary (optimized version)
        success, summary = ProjectService.get_monthly_roster_summary_only(member, today.year, today.month)
        
        if not success:
            summary = {
                'present_days': 0,
                'leave_days': 0,
                'weekoff_days': 0,
                'task_hours': '00:00',
                'misc_hours': '00:00',
                'total_hours': '00:00'
            }
        
        # Get previous day snapshot
        previous_day = previous_day_snapshots.get(member.id, {
            'date': today - timedelta(days=1),
            'status': 'NO_DATA',
            'status_display': 'No Data',
            'total_minutes': 0,
            'total_formatted': '00:00',
            'task_minutes': 0,
            'misc_minutes': 0,
            'tasks': [],
            'misc_activities': [],
            'has_data': False
        })
        
        # Calculate efficiency percentage for previous day
        if previous_day['total_minutes'] > 0:
            # Assuming 8 hours (480 minutes) as standard working day
            efficiency_percentage = (previous_day['total_minutes'] / 480) * 100
        else:
            efficiency_percentage = 0
        
        previous_day['efficiency_percentage'] = round(efficiency_percentage, 1)
        
        # Determine card status color
        if previous_day['status'] == 'PRESENT':
            if previous_day['total_minutes'] >= 420:  # 7+ hours
                card_status_class = 'status-excellent'
            elif previous_day['total_minutes'] >= 360:  # 6+ hours
                card_status_class = 'status-good'
            elif previous_day['total_minutes'] >= 240:  # 4+ hours
                card_status_class = 'status-fair'
            else:
                card_status_class = 'status-low'
        elif previous_day['status'] in ['LEAVE', 'SICK_LEAVE']:
            card_status_class = 'status-leave'
        elif previous_day['status'] == 'WEEK_OFF':
            card_status_class = 'status-weekoff'
        elif previous_day['status'] == 'HOLIDAY':
            card_status_class = 'status-holiday'
        else:
            card_status_class = 'status-nodata'
        
        previous_day['card_status_class'] = card_status_class
        
        team_member_data.append({
            'member': member,
            'summary': summary,
            'previous_day': previous_day
        })
    
    # Sort team members by previous day total hours (most productive first)
    team_member_data.sort(key=lambda x: x['previous_day']['total_minutes'], reverse=True)
    
    context = {
        'team_members': team_member_data,
        'current_month': today.strftime('%B %Y'),
        'previous_date': previous_day_snapshots[team_members[0].id]['date'] if team_members and team_members[0].id in previous_day_snapshots else today - timedelta(days=1),
        'title': 'Team Roster',
        'show_previous_day': True  # Feature flag for easy rollback
    }
    
    return render(request, 'projects/team_roster_list.html', context)


@login_required
def team_member_monthly_roster(request, team_member_id, year=None, month=None):
    """
    Display monthly roster for a specific team member (read-only for management).
    """
    # Check if user has management access (DPM, VIDEO_PM, or SENIOR_MANAGER)
    redirect_response = ensure_has_management_access(request)
    if redirect_response:
        return redirect_response
    
    try:
        team_member = User.objects.get(id=team_member_id, role='TEAM_MEMBER')
    except User.DoesNotExist:
        messages.error(request, "Team member not found.")
        return redirect('projects:team_roster_list')
    
    # Default to current month if not specified
    if not year or not month:
        today = date.today()
        year, month = today.year, today.month
    
    # Get monthly roster data
    success, result = ProjectService.get_monthly_roster(team_member, year, month)
    
    if not success:
        messages.error(request, result)
        return redirect('projects:team_roster_list')
    
    monthly_data = result
    
    # Calculate navigation dates
    current_date = date(year, month, 1)
    prev_month = current_date - timedelta(days=1)
    next_month_day = current_date.replace(day=28) + timedelta(days=4)
    next_month = next_month_day - timedelta(days=next_month_day.day-1)
    
    context = {
        'monthly_data': monthly_data,
        'current_date': current_date,
        'prev_month': prev_month,
        'next_month': next_month,
        'team_member': team_member,
        'is_read_only': True,  # Flag for template to hide edit features
        'title': f'{team_member.get_full_name()} - Roster - {monthly_data["month_name"]} {year}'
    }
    
    return render(request, 'projects/team_member_monthly_roster.html', context)


@login_required
def team_member_daily_roster(request):
    """
    Display daily roster for a specific team member (read-only for management).
    """
    # Check if user has management access (DPM, VIDEO_PM, or SENIOR_MANAGER)
    redirect_response = ensure_has_management_access(request)
    if redirect_response:
        return redirect_response
    
    # Get team member from query params
    team_member_id = request.GET.get('team_member')
    if not team_member_id:
        messages.error(request, "Team member not specified.")
        return redirect('projects:team_roster_list')
    
    try:
        team_member = User.objects.get(id=team_member_id, role='TEAM_MEMBER')
    except User.DoesNotExist:
        messages.error(request, "Team member not found.")
        return redirect('projects:team_roster_list')
    
    # Get filter parameters
    try:
        default_date = timezone.localtime(timezone.now()).date()
        date_str = request.GET.get('date')
        selected_date = datetime.strptime(date_str, '%Y-%m-%d').date() if date_str else default_date
    except (ValueError, TypeError):
        selected_date = default_date
    
    show_week = request.GET.get('week_view') == 'on'
    
    # Get roster data
    success, roster_data = ProjectService.get_daily_roster_data(
        team_member, selected_date, show_week
    )
    
    if not success:
        messages.error(request, roster_data)
        roster_data = {
            'daily_totals': [],
            'daily_rosters': {},
            'misc_hours_entries': [],
            'date_range': 'Error loading data',
            'total_formatted': '00:00',
            'assignment_minutes': 0,
            'misc_minutes': 0,
        }
    
    # Calculate daily summaries if in week view
    daily_summaries = {}
    if show_week and success:
        day_totals_map = {}
        for dt in roster_data['daily_totals']:
            day_key = dt.date_worked
            day_totals_map.setdefault(day_key, 0)
            day_totals_map[day_key] += dt.total_minutes
        
        for me in roster_data['misc_hours_entries']:
            day_key = me.date
            day_totals_map.setdefault(day_key, 0)
            day_totals_map[day_key] += me.duration_minutes
        
        for day, total_minutes in day_totals_map.items():
            daily_summaries[day] = ProjectService._format_minutes(total_minutes)
    
    # Create filter form for rendering
    filter_form = DailyRosterFilterForm(initial={
        'date': selected_date,
        'week_view': show_week,
    })
    
    context = {
        'daily_totals': roster_data['daily_totals'],
        'daily_rosters': roster_data['daily_rosters'],
        'misc_hours_entries': roster_data['misc_hours_entries'],
        'filter_form': filter_form,
        'selected_date': selected_date,
        'show_week': show_week,
        'date_range': roster_data['date_range'],
        'total_formatted': roster_data['total_formatted'],
        'assignment_minutes': roster_data['assignment_minutes'],
        'misc_minutes': roster_data['misc_minutes'],
        'assignment_minutes_formatted': ProjectService._format_minutes(roster_data['assignment_minutes']),
        'misc_minutes_formatted': ProjectService._format_minutes(roster_data['misc_minutes']),
        'daily_summaries': daily_summaries,
        'team_member': team_member,
        'is_read_only': True,  # Flag for template to hide edit features
        'title': f'{team_member.get_full_name()} - Daily Roster - {roster_data["date_range"]}'
    }
    
    return render(request, 'projects/team_member_daily_roster.html', context)


@login_required
def get_active_timers_json(request):
    """
    AJAX endpoint to fetch active timers data in JSON format.
    Only accessible to management roles (DPM, VIDEO_PM, SENIOR_MANAGER).
    
    Returns:
        JsonResponse: Active timers data and summary
    """
    from django.views.decorators.http import require_GET
    from django.conf import settings
    import logging
    
    logger = logging.getLogger(__name__)
    
    # Check user authorization
    if not hasattr(request.user, 'role'):
        return JsonResponse({'error': 'User role not defined'}, status=403)
    
    if request.user.role not in ['DPM', 'VIDEO_PM', 'SENIOR_MANAGER']:
        return JsonResponse({'error': 'Unauthorized access'}, status=403)
    
    try:
        # Get active timers from service
        active_timers = ProjectService.get_all_active_timers()
        summary = ProjectService.get_active_timers_summary()
        
        # Convert to JSON-serializable format
        timers_data = []
        for timer in active_timers:
            timer_json = {
                'id': str(timer['id']),
                'team_member': {
                    'id': timer['team_member']['id'],
                    'username': timer['team_member']['username'],
                    'full_name': timer['team_member']['full_name'],
                    'initials': timer['team_member']['initials']
                },
                'assignment': {
                    'id': str(timer['assignment']['id']),
                    'assignment_id': timer['assignment']['assignment_id'],
                    'sub_task': timer['assignment']['sub_task'],
                    'expected_delivery_date': timer['assignment']['expected_delivery_date'].isoformat() if timer['assignment']['expected_delivery_date'] else None,
                    'projected_formatted': timer['assignment']['projected_formatted']
                },
                'project': {
                    'id': timer['project']['id'],
                    'hs_id': timer['project']['hs_id'],
                    'name': timer['project']['name'],
                    'product': timer['project']['product']
                },
                'task': {
                    'id': timer['task']['id'],
                    'name': timer['task']['name']
                },
                'timer': {
                    'started_at': timer['timer']['started_at'].isoformat(),
                    'started_at_formatted': timer['timer']['started_at_formatted'],
                    'elapsed_seconds': timer['timer']['elapsed_seconds'],
                    'elapsed_formatted': timer['timer']['elapsed_formatted'],
                    'is_long_running': timer['timer']['is_long_running']
                },
                'progress': {
                    'total_worked_formatted': timer['progress']['total_worked_formatted'],
                    'percentage': timer['progress']['percentage'],
                    'status_class': timer['progress']['status_class']
                },
                'urgency': {
                    'status': timer['urgency']['status'],
                    'days_until_deadline': timer['urgency']['days_until_deadline'],
                    'status_class': timer['urgency']['status_class'],
                    'display_text': timer['urgency']['display_text']
                }
            }
            timers_data.append(timer_json)
        
        response_data = {
            'success': True,
            'timers': timers_data,
            'summary': {
                'count': summary['count'],
                'display_text': summary['display_text'],
                'has_timers': summary['has_timers'],
                'team_members': summary.get('team_members', []),
                'more_count': summary.get('more_count', 0)
            },
            'timestamp': timezone.now().isoformat()
        }
        
        return JsonResponse(response_data)
        
    except Exception as e:
        logger.exception(f"Error in get_active_timers_json: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': 'Failed to fetch active timers',
            'message': str(e) if getattr(settings, 'DEBUG', False) else 'An error occurred'
        }, status=500)






