from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.db import transaction
# Export functionality imports
import csv
from openpyxl import Workbook
from datetime import date, datetime
from django.db.models import Q
from .models import VideoProject, VideoProjectStatusOption, VideoProduct
from .services import VideoProjectService
from .forms import (
    VideoProjectCreateForm, VideoProjectStatusUpdateForm, VideoProjectFilterForm, VideoProjectEditForm
)

def ensure_is_video_pm(user):
    """Helper function to check VIDEO_PM role"""
    if user.role != 'VIDEO_PM':
        raise PermissionDenied("Access denied. Video Production Manager role required.")

@login_required
def video_create_project(request):
    """Create new video production project - VIDEO_PM only"""
    if request.user.role != 'VIDEO_PM':
        messages.error(request, "Access denied. Only Video Production Managers can create projects.")
        return redirect('home')
    
    if request.method == 'POST':
        form = VideoProjectCreateForm(request.POST)
        if form.is_valid():
            success, result = VideoProjectService.create_video_project(
                form.cleaned_data, 
                request.user
            )
            if success:
                project = result
                messages.success(request, f'Video project "{project.project_name}" created successfully with ID: {project.hs_id}')
                return redirect('video_production:project_detail', project_id=project.id)
            else:
                # Handle error - result contains error message or dict
                if isinstance(result, dict):
                    # Field-specific errors
                    for field, errors in result.items():
                        for error in errors:
                            messages.error(request, f'{field}: {error}')
                else:
                    # General error message
                    messages.error(request, f'Error creating project: {result}')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = VideoProjectCreateForm()
    
    return render(request, 'video_production/video_create_project.html', {
        'form': form,
        'page_title': 'Create Video Production Project'
    })

@login_required
def video_project_detail(request, project_id):
    """View video project details - simplified version mirroring projects app"""
    if request.user.role != 'VIDEO_PM':
        messages.error(request, "Access denied. Only Video Production Managers can view project details.")
        return redirect('home')
    
    success, result = VideoProjectService.get_video_project(project_id)
    if not success:
        messages.error(request, result)
        return redirect('video_production:project_list')
    
    project = result
    
    # Check if user owns this project
    if project.video_pm != request.user:
        messages.error(request, "You can only view your own projects.")
        return redirect('video_production:project_list')
    
    # Get comprehensive project details
    success, details = VideoProjectService.get_video_project_details(project_id)
    if not success:
        messages.error(request, details)
        return redirect('video_production:project_list')
    
    # Get status options for the update modal
    filter_options = VideoProjectService.get_video_filter_options()
    
    return render(request, 'video_production/video_project_detail.html', {
        'project': details['project'],
        'status_history': details['status_history'],
        'deliveries': details['deliveries'],
        'status_options': filter_options['statuses'],
        'page_title': f'Video Project - {project.hs_id}'
    })

@login_required
@require_http_methods(["POST"])
def video_update_project_status(request, project_id):
    """Update project status - supports AJAX"""
    # Debug: Check if this is an AJAX request
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    
    if request.user.role != 'VIDEO_PM':
        if is_ajax:
            return JsonResponse({'success': False, 'error': 'Access denied'})
        else:
            messages.error(request, 'Access denied')
            return redirect('video_production:project_detail', project_id=project_id)
    
    success, result = VideoProjectService.get_video_project(project_id)
    if not success:
        if is_ajax:
            return JsonResponse({'success': False, 'error': result})
        else:
            messages.error(request, result)
            return redirect('video_production:project_detail', project_id=project_id)
    
    project = result
    
    # Check if user owns this project
    if project.video_pm != request.user:
        if is_ajax:
            return JsonResponse({'success': False, 'error': 'You can only update your own projects'})
        else:
            messages.error(request, 'You can only update your own projects')
            return redirect('video_production:project_detail', project_id=project_id)
    
    form = VideoProjectStatusUpdateForm(request.POST)
    if form.is_valid():
        try:
            updated_project = VideoProjectService.update_project_status(
                project_id,
                form.cleaned_data['status'].id,
                form.cleaned_data['comments'],
                request.user
            )
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': f'Status updated to {updated_project.current_status.name}',
                    'new_status': updated_project.current_status.name
                })
            else:
                messages.success(request, f'Project status updated to {updated_project.current_status.name}')
                return redirect('video_production:project_detail', project_id=project_id)
        except Exception as e:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': f'Error updating status: {str(e)}'})
            else:
                messages.error(request, f'Error updating status: {str(e)}')
                return redirect('video_production:project_detail', project_id=project_id)
    else:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'errors': form.errors})
        else:
            messages.error(request, 'Error updating status. Please check your input.')
            return redirect('video_production:project_detail', project_id=project_id)

@login_required
def video_project_list(request):
    """List pipeline (active) video projects - mirrors projects app structure"""
    if request.user.role != 'VIDEO_PM':
        messages.error(request, "Access denied. Only Video Production Managers can view projects.")
        return redirect('home')
    
    # Get filter form
    filter_form = VideoProjectFilterForm(request.GET or None)
    
    # Build filters dict from form data
    filters = {}
    filters_applied = {}
    filters_applied_display = {}
    
    if filter_form.is_valid():
        for field_name, value in filter_form.cleaned_data.items():
            if value:
                filters[field_name] = value
                filters_applied[field_name] = value
                
                # Get display names for filters (similar to projects app)
                if field_name == 'status' and hasattr(value, 'name'):
                    filters_applied_display[field_name] = value.name
                elif field_name == 'product' and hasattr(value, 'name'):
                    filters_applied_display[field_name] = value.name
                elif field_name == 'region' and hasattr(value, 'name'):
                    filters_applied_display[field_name] = value.name
                elif field_name == 'city' and hasattr(value, 'name'):
                    filters_applied_display[field_name] = value.name
                elif field_name == 'video_pm' and hasattr(value, 'get_full_name'):
                    filters_applied_display[field_name] = value.get_full_name()
                else:
                    filters_applied_display[field_name] = str(value)
    
    # Get projects based on filters - always use pipeline project type for this view
    success, projects_queryset = VideoProjectService.get_video_project_list(request.user, filters, project_type='pipeline')
    if not success:
        messages.error(request, f"Error loading projects: {projects_queryset}")
        projects_queryset = VideoProject.objects.none()
    
    # Pagination
    paginator = Paginator(projects_queryset, 25)  # Match projects app pagination
    page_number = request.GET.get('page')
    projects = paginator.get_page(page_number)
    
    # Add latest_status_date to each project (for template consistency)
    for project in projects:
        # Get the latest status history date
        latest_history = project.status_history.order_by('-changed_at').first()
        project.latest_status_date = latest_history.changed_at if latest_history else project.created_at
    
    return render(request, 'video_production/video_project_list.html', {
        'title': 'Video Production Pipeline',
        'projects': projects,
        'filter_form': filter_form,
        'filters_applied': filters_applied,
        'filters_applied_display': filters_applied_display,
        'is_pipeline': True,
        'page_title': 'Video Production Pipeline',
        'show_pipeline': True
    })

@login_required
def video_delivered_projects(request):
    """List delivered video projects - mirrors projects app structure"""
    if request.user.role != 'VIDEO_PM':
        messages.error(request, "Access denied. Only Video Production Managers can view projects.")
        return redirect('home')
    
    # Get filter form
    filter_form = VideoProjectFilterForm(request.GET or None)
    
    # Build filters dict from form data
    filters = {}
    filters_applied = {}
    filters_applied_display = {}
    
    if filter_form.is_valid():
        for field_name, value in filter_form.cleaned_data.items():
            if value:
                filters[field_name] = value
                filters_applied[field_name] = value
                
                # Get display names for filters (similar to projects app)
                if field_name == 'status' and hasattr(value, 'name'):
                    filters_applied_display[field_name] = value.name
                elif field_name == 'product' and hasattr(value, 'name'):
                    filters_applied_display[field_name] = value.name
                elif field_name == 'region' and hasattr(value, 'name'):
                    filters_applied_display[field_name] = value.name
                elif field_name == 'city' and hasattr(value, 'name'):
                    filters_applied_display[field_name] = value.name
                elif field_name == 'video_pm' and hasattr(value, 'get_full_name'):
                    filters_applied_display[field_name] = value.get_full_name()
                else:
                    filters_applied_display[field_name] = str(value)
    
    # Get projects based on filters - always use delivered project type for this view
    success, projects_queryset = VideoProjectService.get_video_project_list(request.user, filters, project_type='delivered')
    if not success:
        messages.error(request, f"Error loading projects: {projects_queryset}")
        projects_queryset = VideoProject.objects.none()
    
    # Pagination
    paginator = Paginator(projects_queryset, 25)  # Match projects app pagination
    page_number = request.GET.get('page')
    projects = paginator.get_page(page_number)
    
    # Add delivery_date to each project (for template consistency)
    for project in projects:
        # Use the property that gets delivery date from status history
        project.latest_status_date = project.delivery_date or project.created_at
    
    return render(request, 'video_production/video_project_list.html', {
        'title': 'Delivered Video Projects',
        'projects': projects,
        'filter_form': filter_form,
        'filters_applied': filters_applied,
        'filters_applied_display': filters_applied_display,
        'is_pipeline': False,
        'page_title': 'Delivered Video Projects',
        'show_delivered': True
    })

@login_required
def export_pipeline_video_projects(request):
    """
    Export pipeline video projects (non-delivered) to CSV or XLSX format.
    Uses the same filtering logic as video_project_list view.
    Only accessible by VIDEO_PMs.
    """
    # Check if user is a VIDEO_PM
    if request.user.role != 'VIDEO_PM':
        messages.error(request, "Access denied. This page is only for Video Production Managers.")
        return redirect('home')
    
    # Get export format from request (default to CSV)
    export_format = request.GET.get('format', 'csv').lower()
    
    # Get filter parameters from request (same as video_project_list view)
    search_query = request.GET.get('search', '')
    status = request.GET.get('status', '')
    product = request.GET.get('product', '')
    region = request.GET.get('region', '')
    city = request.GET.get('city', '')
    video_pm = request.GET.get('video_pm', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    # Convert date strings to date objects if provided
    date_from_obj = None
    date_to_obj = None
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
        except ValueError:
            pass
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
        except ValueError:
            pass

    # Prepare filters dictionary
    filters = {}
    if search_query:
        filters['search'] = search_query
    if status:
        try:
            status_obj = VideoProjectStatusOption.objects.get(id=status)
            filters['status'] = status_obj
        except VideoProjectStatusOption.DoesNotExist:
            pass
    if product:
        try:
            product_obj = VideoProduct.objects.get(id=product)
            filters['product'] = product_obj
        except VideoProduct.DoesNotExist:
            pass
    if region:
        try:
            from locations.models import Region
            region_obj = Region.objects.get(id=region)
            filters['region'] = region_obj
        except Region.DoesNotExist:
            pass
    if city:
        try:
            from locations.models import City
            city_obj = City.objects.get(id=city)
            filters['city'] = city_obj
        except City.DoesNotExist:
            pass
    if video_pm:
        try:
            from accounts.models import User
            video_pm_obj = User.objects.get(id=video_pm)
            filters['video_pm'] = video_pm_obj
        except User.DoesNotExist:
            pass
    if date_from_obj:
        filters['date_from'] = date_from_obj
    if date_to_obj:
        filters['date_to'] = date_to_obj

    # Get ALL pipeline video projects (no pagination for export)
    success, projects_queryset = VideoProjectService.get_video_project_list(
        request.user, filters, project_type='pipeline'
    )

    if not success:
        messages.error(request, f"Error exporting video projects: {projects_queryset}")
        return redirect('video_production:project_list')

    # Convert queryset to list
    projects_list = list(projects_queryset)

    # Define the columns for export
    headers = [
        'HS ID', 'Opportunity ID', 'Project Name', 'Builder Name', 'City', 'Region',
        'Product', 'Package ID', 'Quantity', 'Purchase Date',
        'Sales Confirmation Date', 'Expected TAT', 'Account Manager', 'Video PM',
        'Current Status', 'Expected Completion Date', 'Created At'
    ]

    if export_format == 'xlsx':
        return _export_video_projects_to_xlsx(projects_list, headers)
    else:
        return _export_video_projects_to_csv(projects_list, headers)


def _export_video_projects_to_csv(projects, headers):
    """Helper function to export video projects to CSV format."""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="pipeline_video_projects_{date.today().strftime("%Y%m%d")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(headers)
    
    for project in projects:
        row = [
            project.hs_id or '',
            project.opportunity_id or '',
            project.project_name or '',
            project.builder_name or '',
            project.city.name if project.city else '',
            project.city.region.name if project.city and project.city.region else '',
            project.product.name if project.product else '',
            project.package_id or '',
            project.quantity or '',
            project.purchase_date.strftime('%Y-%m-%d') if project.purchase_date else '',
            project.sales_confirmation_date.strftime('%Y-%m-%d') if project.sales_confirmation_date else '',
            project.expected_tat or '',
            project.account_manager or '',
            project.video_pm.get_full_name() if project.video_pm else '',
            project.current_status.name if project.current_status else '',
            project.expected_completion_date.strftime('%Y-%m-%d') if project.expected_completion_date else '',
            project.created_at.strftime('%Y-%m-%d %H:%M:%S') if project.created_at else ''
        ]
        writer.writerow(row)
    
    return response


def _export_video_projects_to_xlsx(projects, headers):
    """Helper function to export video projects to XLSX format."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Pipeline Video Projects"
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)
    
    # Write data
    for row_num, project in enumerate(projects, 2):
        data = [
            project.hs_id or '',
            project.opportunity_id or '',
            project.project_name or '',
            project.builder_name or '',
            project.city.name if project.city else '',
            project.city.region.name if project.city and project.city.region else '',
            project.product.name if project.product else '',
            project.package_id or '',
            project.quantity or '',
            project.purchase_date.strftime('%Y-%m-%d') if project.purchase_date else '',
            project.sales_confirmation_date.strftime('%Y-%m-%d') if project.sales_confirmation_date else '',
            project.expected_tat or '',
            project.account_manager or '',
            project.video_pm.get_full_name() if project.video_pm else '',
            project.current_status.name if project.current_status else '',
            project.expected_completion_date.strftime('%Y-%m-%d') if project.expected_completion_date else '',
            project.created_at.strftime('%Y-%m-%d %H:%M:%S') if project.created_at else ''
        ]
        
        for col_num, value in enumerate(data, 1):
            worksheet.cell(row=row_num, column=col_num, value=value)
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="pipeline_video_projects_{date.today().strftime("%Y%m%d")}.xlsx"'
    
    workbook.save(response)
    return response

@login_required
def video_edit_project(request, project_id):
    """Edit existing video project"""
    if request.user.role != 'VIDEO_PM':
        messages.error(request, "Access denied. Only Video Production Managers can edit projects.")
        return redirect('home')
    
    try:
        project = VideoProjectService.get_video_project(project_id)
        
        # Check if user owns this project
        if project.video_pm != request.user:
            messages.error(request, "You can only edit your own projects.")
            return redirect('video_production:project_list')
        
        if request.method == 'POST':
            form = VideoProjectEditForm(request.POST, instance=project)
            if form.is_valid():
                form.save()
                messages.success(request, f'Project {project.hs_id} updated successfully')
                return redirect('video_production:project_detail', project_id=project_id)
            else:
                messages.error(request, 'Please correct the errors below.')
        else:
            form = VideoProjectEditForm(instance=project)
        
        return render(request, 'video_production/video_edit_project.html', {
            'form': form,
            'project': project,
            'page_title': f'Edit Project - {project.hs_id}'
        })
    
    except VideoProject.DoesNotExist:
        messages.error(request, 'Video project not found.')
        return redirect('video_production:project_list')

@login_required
def video_complete_project(request, project_id):
    """Mark project as completed and track delivery performance"""
    if request.user.role != 'VIDEO_PM':
        messages.error(request, "Access denied. Only Video Production Managers can complete projects.")
        return redirect('home')
    
    try:
        project = VideoProjectService.get_video_project(project_id)
        
        # Check if user owns this project
        if project.video_pm != request.user:
            messages.error(request, "You can only complete your own projects.")
            return redirect('video_production:project_list')
        
        delivery = VideoProjectService.track_video_project_delivery(project_id)
        messages.success(request, f'Project {project.hs_id} marked as completed. Delivery performance: {delivery.delivery_performance_rating}')
        
        return redirect('video_production:project_detail', project_id=project_id)
    
    except VideoProject.DoesNotExist:
        messages.error(request, 'Video project not found.')
        return redirect('video_production:project_list')

def ensure_has_management_access(user):
    """Helper function to check management access for reports (DPM, VIDEO_PM, SENIOR_MANAGER)"""
    allowed_roles = ['DPM', 'VIDEO_PM', 'SENIOR_MANAGER']
    if user.role not in allowed_roles:
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied("Access denied. Management role required.")

@login_required
def video_report(request):
    """
    Video Production Report for Senior Managers - similar to general_report.
    Shows Sales Confirmed, 1st Cut Deliveries, and Final Deliveries metrics.
    Accessible by SENIOR_MANAGER, VIDEO_PM, and DPM roles.
    """
    # Check management access
    try:
        ensure_has_management_access(request.user)
    except Exception:
        messages.error(request, "Access denied. Management role required.")
        return redirect('home')
    
    # Parse filters from GET request
    filters = {}
    if request.GET.get('date_from'):
        try:
            filters['date_from'] = datetime.strptime(request.GET['date_from'], '%Y-%m-%d').date()
        except ValueError:
            pass
    else:
        # Default to July 1, 2025 if no date_from provided
        from datetime import date
        filters['date_from'] = date(2025, 7, 1)
    
    if request.GET.get('date_to'):
        try:
            filters['date_to'] = datetime.strptime(request.GET['date_to'], '%Y-%m-%d').date()
        except ValueError:
            pass
    
    if request.GET.get('product'):
        filters['product'] = request.GET['product']
    
    if request.GET.get('video_pm'):
        filters['video_pm'] = request.GET['video_pm']
    
    # Get report data from service
    report_data = VideoProjectService.get_video_report_data(filters)
    
    context = {
        'sales_confirmed': report_data['sales_confirmed'],
        'first_cut_deliveries': report_data['first_cut_deliveries'],
        'final_deliveries': report_data['final_deliveries'],
        'filters': report_data['filters'],
        'products': report_data['products'],
        'video_pms': report_data['video_pms'],
        'product_chart_json': report_data['product_chart_json'],
        'video_pm_chart_json': report_data['video_pm_chart_json'],
        'fcd_product_chart_json': report_data['fcd_product_chart_json'],
        'fcd_video_pm_chart_json': report_data['fcd_video_pm_chart_json'],
        'fd_product_chart_json': report_data['fd_product_chart_json'],
        'fd_video_pm_chart_json': report_data['fd_video_pm_chart_json'],
        'page_title': 'Video Production Report'
    }
    
    return render(request, 'video_production/reports/video_report.html', context)
